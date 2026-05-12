import csv
from datetime import date, datetime
from io import BytesIO

from django.core.files.base import ContentFile
from django.db.models import Q, F
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.pacientes.exports import crear_excel_pacientes, excel_response
from apps.pacientes.models import MovimientoPaciente, Paciente
from apps.pacientes.services import categoria_por_diagnostico, prioridad_normalizada
from apps.usuarios.permissions import IsAdminOrAdministrativoRole

from .models import ImportacionMensual
from .parser import normalizar_texto, parsear_derivaciones, parsear_fecha, previsualizar_derivaciones
from .serializers import ImportacionDerivacionesSerializer


MESES_SHEET = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

MESES_LABEL = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

ESTADOS_EGRESADOS = [
    Paciente.Estado.ALTA_MEDICA,
    Paciente.Estado.EGRESO_VOLUNTARIO,
    Paciente.Estado.EGRESO_ADMINISTRATIVO,
    Paciente.Estado.ABANDONO,
    Paciente.Estado.DERIVADO,
]


def _bool_from_request(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "on", "si"}


def _mes_y_anio_referencia(importacion: ImportacionMensual) -> tuple[int, int]:
    return (
        importacion.mes_datos or importacion.mes,
        importacion.anio_datos or importacion.anio,
    )


def _periodo_q(mes: int, anio: int) -> Q:
    return Q(mes_datos=mes, anio_datos=anio) | Q(
        mes_datos__isnull=True,
        anio_datos__isnull=True,
        mes=mes,
        anio=anio,
    )


def _serialize_importacion(importacion: ImportacionMensual) -> dict:
    mes_ref, anio_ref = _mes_y_anio_referencia(importacion)
    observaciones = [
        observacion
        for observacion in (importacion.observaciones_revision or [])
        if not observacion.get("tipo") or observacion.get("tipo") == "ERROR"
    ]
    return {
        "id": importacion.id,
        "archivo_nombre": importacion.archivo_nombre or importacion.archivo.name.rsplit("/", 1)[-1],
        "mes": importacion.mes,
        "anio": importacion.anio,
        "mes_datos": importacion.mes_datos,
        "anio_datos": importacion.anio_datos,
        "mes_label": MESES_LABEL.get(mes_ref, str(mes_ref)),
        "periodo_label": f"{MESES_LABEL.get(mes_ref, mes_ref)} {anio_ref}",
        "usuario_id": importacion.usuario_id,
        "usuario_nombre": importacion.usuario.nombre if importacion.usuario else None,
        "fecha_subida": importacion.fecha_subida.isoformat(),
        "estado": importacion.estado,
        "estado_label": importacion.get_estado_display(),
        "total_registros": importacion.total_registros,
        "registros_importados": importacion.registros_importados,
        "duplicados": importacion.duplicados,
        "errores": importacion.errores,
        "errores_count": len(importacion.errores or []),
        "observaciones_revision": observaciones,
        "observaciones_revision_count": len(observaciones),
        "observaciones_pendientes_count": len(
            [
                item
                for item in observaciones
                if item.get("estado_revision", "PENDIENTE") == "PENDIENTE"
            ]
        ),
        "pacientes_actuales_del_corte": _resumen_pacientes_corte(mes_ref, anio_ref),
        "reemplazada_por": importacion.reemplazada_por_id,
    }


def _resumen_pacientes_corte(mes: int, anio: int) -> dict:
    pacientes = Paciente.objects.filter(fecha_derivacion__month=mes, fecha_derivacion__year=anio)
    return {
        "pendientes": pacientes.filter(estado=Paciente.Estado.PENDIENTE).count(),
        "rescate": pacientes.filter(estado=Paciente.Estado.RESCATE).count(),
        "ingresados": pacientes.filter(estado=Paciente.Estado.INGRESADO).count(),
        "egresados_total": pacientes.filter(estado__in=ESTADOS_EGRESADOS).count(),
    }


def _id_ccr_desde_pk(pk: int) -> str:
    return f"CCR-{pk:04d}"


def _vaciar_importaciones(importaciones: list[ImportacionMensual], *, borrar_importacion: bool) -> dict:
    # Acción destructiva de mantenimiento: solo borra pacientes importados sin asignación activa.
    # TODO: evaluar archivado lógico de cortes para conservar trazabilidad sin borrar datos físicos.
    pacientes_eliminados = 0
    importaciones_eliminadas = 0
    archivos_eliminados = 0

    for importacion in importaciones:
        pacientes_qs = importacion.pacientes_creados.filter(
            kine_asignado__isnull=True,
            estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE],
        )
        pacientes_eliminados += pacientes_qs.count()
        pacientes_qs.delete()

        if borrar_importacion:
            if importacion.archivo:
                importacion.archivo.delete(save=False)
                archivos_eliminados += 1
            importacion.delete()
            importaciones_eliminadas += 1

    return {
        "pacientes_eliminados": pacientes_eliminados,
        "importaciones_eliminadas": importaciones_eliminadas,
        "archivos_eliminados": archivos_eliminados,
    }


def _normalizar_rut(rut: str | None) -> str:
    return (rut or "").replace(".", "").replace("-", "").upper().strip()


ENCABEZADO_DERIVACIONES_PEGADO = [
    [
        "FECHA",
        "",
        "",
        "",
        "",
        "DIAGNÓSTICO MÉDICO",
        "PROFESIONAL DERIVADO",
        "GRADO PRIORIDAD",
        "OBSERVACIONES",
        "USUARIO PREFERENTE",
        "",
        "",
        "OBJETIVOS DEL TRATAMIENTO (OBSERVACIONES)",
    ],
    [
        "",
        "NOMBRE",
        "RUT",
        "EDAD",
        "DESDE",
        "",
        "",
        "",
        "",
        "MAYOR O IGUAL 60",
        "DISCAPACIDAD",
        "CUIDADOR/RA",
        "",
    ],
]

CENTROS_DESDE_PEGADO = {"CAR", "CES", "CCE", "HT", "HH", "FST", "FST HT", "TMT", "TMT HT"}
PRIORIDADES_PEGADO = {"ALTA", "MEDIANA", "MEDIA", "MODERADA", "BAJA"}
PROFESIONALES_PEGADO = {
    "KINESIOLOGO",
    "KINESIOLOGA",
    "KINESIOLOGIA",
    "KINESIOLGO",
    "FONOAUDIOLOGIA",
    "TERAPIA OCUPACIONAL",
    "TERAPEUTA OCUPACIONAL",
}


def _fila_contiene_encabezado_pegado(fila: list[str]) -> bool:
    normalizadas = {normalizar_texto(celda) for celda in fila if str(celda).strip()}
    if not normalizadas:
        return False
    encabezados_fuertes = {
        "FECHA",
        "FECHA DERIVACION",
        "DIAGNOSTICO MEDICO",
        "PROFESIONAL DERIVADO",
        "GRADO PRIORIDAD",
        "USUARIO PREFERENTE",
    }
    encabezados_paciente = {"NOMBRE", "RUT", "EDAD", "DESDE"}
    return bool(normalizadas & encabezados_fuertes) or len(normalizadas & encabezados_paciente) >= 2


def _celda_parece_rut_pegado(valor: str) -> bool:
    if parsear_fecha(valor):
        return False
    rut = _normalizar_rut(valor)
    return len(rut) >= 7 and rut[:-1].isdigit() and (rut[-1].isdigit() or rut[-1] == "K")


def _celda_parece_edad_pegado(valor: str) -> bool:
    try:
        edad = int(str(valor).strip())
    except (TypeError, ValueError):
        return False
    return 0 < edad < 120


def _fila_parece_dato_pegado(fila: list[str]) -> bool:
    if not fila or parsear_fecha(fila[0]) is None:
        return False
    rut_en_primeras_columnas = any(_celda_parece_rut_pegado(celda) for celda in fila[1:5])
    nombre_en_primeras_columnas = any(str(celda).strip() for celda in fila[1:3])
    return rut_en_primeras_columnas and nombre_en_primeras_columnas


def _filas_tienen_encabezado_pegado(filas: list[list[str]]) -> bool:
    return any(_fila_contiene_encabezado_pegado(fila) for fila in filas[:3])


def _fila_basura_pegado(fila: list[str]) -> bool:
    no_vacias = [celda for celda in fila if str(celda).strip()]
    return len(no_vacias) == 1 and parsear_fecha(no_vacias[0]) is not None


def _normalizar_fila_datos_pegada(fila: list[str]) -> list[str]:
    if not _fila_parece_dato_pegado(fila):
        return fila

    fecha_idx = next((idx for idx, celda in enumerate(fila[:3]) if parsear_fecha(celda)), 0)
    rut_idx = next((idx for idx, celda in enumerate(fila) if _celda_parece_rut_pegado(celda)), -1)
    edad_idx = next(
        (
            idx
            for idx, celda in enumerate(fila)
            if rut_idx < idx <= rut_idx + 3 and _celda_parece_edad_pegado(celda)
        ),
        -1,
    )
    desde_idx = next(
        (
            idx
            for idx, celda in enumerate(fila)
            if edad_idx < idx <= edad_idx + 4 and normalizar_texto(celda) in CENTROS_DESDE_PEGADO
        ),
        -1,
    )
    profesional_idx = next(
        (
            idx
            for idx, celda in enumerate(fila)
            if idx > max(edad_idx, desde_idx) and normalizar_texto(celda) in PROFESIONALES_PEGADO
        ),
        -1,
    )
    prioridad_idx = next(
        (
            idx
            for idx, celda in enumerate(fila)
            if idx > profesional_idx and normalizar_texto(celda) in PRIORIDADES_PEGADO
        ),
        -1,
    )

    if rut_idx < 0 or edad_idx < 0 or profesional_idx < 0 or prioridad_idx < 0:
        return fila

    nombre = " ".join(celda for celda in fila[fecha_idx + 1:rut_idx] if str(celda).strip())
    diagnostico_inicio = desde_idx + 1 if desde_idx >= 0 else edad_idx + 1
    diagnostico = " ".join(celda for celda in fila[diagnostico_inicio:profesional_idx] if str(celda).strip())
    resto = fila[prioridad_idx + 1:]

    return [
        fila[fecha_idx] if fecha_idx < len(fila) else "",
        nombre or (fila[1] if len(fila) > 1 else ""),
        fila[rut_idx],
        fila[edad_idx],
        fila[desde_idx] if desde_idx >= 0 else "",
        diagnostico,
        fila[profesional_idx],
        fila[prioridad_idx],
        resto[0] if len(resto) > 0 else "",
        resto[1] if len(resto) > 1 else "",
        resto[2] if len(resto) > 2 else "",
        resto[3] if len(resto) > 3 else "",
        " ".join(celda for celda in resto[4:] if str(celda).strip()),
    ]


def _mes_desde_fecha_preview(valor: str | None) -> int | None:
    fecha = _fecha_desde_preview(valor)
    return fecha.month if fecha else None


def _fecha_desde_preview(valor: str | None) -> date | None:
    if not valor:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(valor).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _periodo_desde_fecha_preview(valor: str | None) -> tuple[int, int] | None:
    fecha = _fecha_desde_preview(valor)
    return (fecha.month, fecha.year) if fecha else None


def _mes_desde_registro(registro: dict, mes_fallback: int | None = None) -> int | None:
    hoja = str(registro.get("hoja") or "").upper()
    for nombre, numero in MESES_SHEET.items():
        if nombre in hoja:
            return numero
    return _mes_desde_fecha_preview(registro.get("fecha_derivacion")) or mes_fallback


def _periodo_desde_registro(
    registro: dict,
    mes_fallback: int | None = None,
    anio_fallback: int | None = None,
) -> tuple[int, int] | None:
    periodo = _periodo_desde_fecha_preview(registro.get("fecha_derivacion"))
    if periodo:
        return periodo
    periodo = _periodo_desde_fecha_preview(registro.get("fecha_original"))
    if periodo:
        return periodo
    mes = _mes_desde_registro(registro, mes_fallback)
    if mes and anio_fallback:
        return (mes, anio_fallback)
    return None


def _periodos_desde_resultado(
    resultado: dict,
    *,
    anio_fallback: int,
    mes_fallback: int | None,
) -> dict[tuple[int, int], int]:
    periodos: dict[tuple[int, int], int] = {}
    for raw_key, count in (resultado.get("periodos_detectados") or {}).items():
        try:
            anio_txt, mes_txt = str(raw_key).split("-", 1)
            mes = int(mes_txt)
            anio = int(anio_txt)
        except (TypeError, ValueError):
            continue
        if 1 <= mes <= 12 and count > 0:
            periodos[(mes, anio)] = periodos.get((mes, anio), 0) + int(count)

    if periodos:
        return periodos

    for nombre_hoja, count in (resultado.get("meses_detectados") or {}).items():
        mes = MESES_SHEET.get(str(nombre_hoja).upper())
        if mes and count > 0:
            periodos[(mes, anio_fallback)] = periodos.get((mes, anio_fallback), 0) + int(count)

    if not periodos and mes_fallback and resultado.get("total", 0) > 0:
        periodos[(mes_fallback, anio_fallback)] = int(resultado.get("total", 0))

    return periodos


def _agrupar_registros_por_periodo(
    registros: list[dict],
    *,
    periodos: dict[tuple[int, int], int],
    mes_fallback: int | None,
    anio_fallback: int,
) -> dict[tuple[int, int], list[dict]]:
    agrupados: dict[tuple[int, int], list[dict]] = {periodo: [] for periodo in periodos}
    periodo_default = next(iter(periodos), None)
    for registro in registros:
        periodo = _periodo_desde_registro(registro, mes_fallback, anio_fallback) or periodo_default
        if periodo in agrupados:
            agrupados[periodo].append(registro)
    return agrupados


def _paciente_vinculado_por_registro(registro: dict) -> Paciente | None:
    rut = _normalizar_rut(registro.get("rut"))
    if not rut:
        return None
    return (
        Paciente.objects.select_related("kine_asignado")
        .filter(rut=rut)
        .order_by("-fecha_derivacion", "-id")
        .first()
    )


def _build_observacion_revision(
    *,
    registro: dict,
    tipo: str,
    importacion: ImportacionMensual,
    paciente: Paciente | None,
) -> dict:
    motivo = registro.get("error") or registro.get("motivo") or ""
    if tipo == "RECURRENTE":
        accion = (
            "El paciente ya existía en lista de espera. Se mantiene su ficha operativa y "
            "se registra la aparición de este corte para seguimiento."
        )
        motivo = motivo or "Registro repetido en el sistema."
        tipo_label = "Recurrente"
    elif tipo == "ADVERTENCIA":
        accion = (
            "El paciente fue importado, pero requiere revisión operativa por datos "
            "históricos o asignación no confirmada."
        )
        motivo = motivo or registro.get("motivo_revision") or "Registro requiere revisión operativa."
        tipo_label = "Revisión"
    else:
        accion = (
            "No se creó una ficha operativa nueva porque el registro trae datos incompletos "
            "o inconsistentes. Si el RUT coincide con una ficha operativa existente, queda vinculado para revisión."
        )
        tipo_label = "Error de datos"

    return {
        "tipo": tipo,
        "tipo_label": tipo_label,
        "importacion_id": importacion.id,
        "periodo_label": f"{MESES_LABEL.get(importacion.mes_datos or importacion.mes, importacion.mes)} {importacion.anio_datos or importacion.anio}",
        "archivo_nombre": importacion.archivo_nombre or importacion.archivo.name.rsplit("/", 1)[-1],
        "hoja": registro.get("hoja") or "",
        "fila": registro.get("fila"),
        "motivo": motivo,
        "accion": accion,
        "nombre": registro.get("nombre") or "",
        "rut": _normalizar_rut(registro.get("rut")),
        "fecha_derivacion": registro.get("fecha_derivacion") or "",
        "fecha_original": registro.get("fecha_original") or "",
        "edad": registro.get("edad") or 0,
        "diagnostico": registro.get("diagnostico") or "",
        "prioridad": registro.get("prioridad") or "",
        "percapita_desde": registro.get("percapita_desde") or "",
        "sector_oficial": registro.get("sector_oficial") or "",
        "sector_cesfam": registro.get("sector_cesfam") or "",
        "profesional": registro.get("profesional") or "",
        "categoria": registro.get("categoria") or "",
        "observaciones": registro.get("observaciones") or "",
        "kine_detectado": registro.get("kine_asignado") or "",
        "estado_sugerido": registro.get("estado_sugerido") or "",
        "recepcion_original": registro.get("recepcion_original") or "",
        "asignado_historico": bool(registro.get("asignado_historico")),
        "paciente_id": paciente.id if paciente else None,
        "paciente_rut": paciente.rut if paciente else None,
        "paciente_nombre": paciente.nombre if paciente else None,
        "paciente_estado": paciente.estado if paciente else None,
        "paciente_id_ccr": paciente.id_ccr if paciente else None,
        "kine_asignado_nombre": paciente.kine_asignado.nombre if paciente and paciente.kine_asignado else None,
        "requiere_revision": tipo == "ADVERTENCIA" or (tipo == "ERROR" and paciente is None),
        "estado_revision": registro.get("estado_revision") or "PENDIENTE",
        "resolucion": registro.get("resolucion") or "",
        "resuelto_en": registro.get("resuelto_en"),
        "resuelto_por_id": registro.get("resuelto_por_id"),
        "resuelto_por_nombre": registro.get("resuelto_por_nombre"),
    }


def _registrar_observaciones_revision(
    *,
    registros: list[dict],
    importaciones: dict[tuple[int, int], ImportacionMensual],
    mes_fallback: int | None,
    anio_fallback: int,
    usuario,
) -> None:
    if not registros or not importaciones:
        return

    observaciones_por_importacion: dict[int, list[dict]] = {
        importacion.id: [] for importacion in importaciones.values()
    }
    movimientos_error: list[MovimientoPaciente] = []
    movimientos_error_keys: set[tuple[int, str, int | None]] = set()

    tipos_revision = {"ERROR", "ADVERTENCIA", "RECURRENTE"}

    for registro in registros:
        estado_registro = registro.get("estado")
        tipo = registro.get("tipo_revision") or ("ERROR" if estado_registro == "ERROR" else "")
        if tipo not in tipos_revision:
            continue

        periodo_registro = _periodo_desde_registro(registro, mes_fallback, anio_fallback)
        importacion = (
            importaciones.get(periodo_registro)
            if periodo_registro is not None
            else None
        ) or next(iter(importaciones.values()))
        paciente = _paciente_vinculado_por_registro(registro)
        observacion = _build_observacion_revision(
            registro=registro,
            tipo=tipo,
            importacion=importacion,
            paciente=paciente,
        )
        observaciones_por_importacion.setdefault(importacion.id, []).append(observacion)

        if tipo == "ERROR" and paciente:
            key = (paciente.id, str(registro.get("hoja") or ""), registro.get("fila"))
            if key in movimientos_error_keys:
                continue
            movimientos_error_keys.add(key)
            motivo = observacion["motivo"]
            movimientos_error.append(
                MovimientoPaciente(
                    paciente=paciente,
                    usuario=usuario,
                    estado_anterior=None,
                    estado_nuevo=paciente.estado,
                    notas=(
                        f"Observación del corte {observacion['periodo_label']}: "
                        f"se mantiene ficha operativa existente para revisión. Motivo: {motivo}"
                    ),
                )
            )

    for importacion in importaciones.values():
        importacion.observaciones_revision = observaciones_por_importacion.get(importacion.id, [])
        importacion.save(update_fields=["observaciones_revision"])

    if movimientos_error:
        MovimientoPaciente.objects.bulk_create(movimientos_error, batch_size=200)


def _observaciones_revision_persistidas(importacion: ImportacionMensual) -> list[dict]:
    observaciones = list(importacion.observaciones_revision or [])
    if observaciones:
        normalizadas = []
        changed = False
        for observacion in observaciones:
            if "estado_revision" not in observacion:
                observacion = {**observacion, "estado_revision": "PENDIENTE"}
                changed = True
            normalizadas.append(observacion)
        if changed:
            importacion.observaciones_revision = normalizadas
            importacion.save(update_fields=["observaciones_revision"])
        return normalizadas

    if not importacion.errores:
        return []

    observaciones = [
        _build_observacion_revision(
            registro={**error, "estado": "ERROR"},
            tipo="ERROR",
            importacion=importacion,
            paciente=_paciente_vinculado_por_registro(error),
        )
        for error in importacion.errores
    ]
    importacion.observaciones_revision = observaciones
    importacion.save(update_fields=["observaciones_revision"])
    return observaciones


def _resolve_observacion(importacion: ImportacionMensual, index: int) -> tuple[list[dict], dict] | None:
    observaciones = _observaciones_revision_persistidas(importacion)
    if index < 0 or index >= len(observaciones):
        return None
    return observaciones, observaciones[index]


def _parse_fecha_revisada(valor: str | None):
    fecha = parsear_fecha(valor)
    if fecha is None:
        raise ValueError("Debe ingresar una fecha de derivación válida.")
    return fecha


def _actualizar_observacion_resuelta(
    *,
    observacion: dict,
    accion: str,
    usuario,
    resolucion: str,
    paciente: Paciente | None = None,
) -> dict:
    actualizada = {
        **observacion,
        "estado_revision": "DESCARTADO" if accion == "DESCARTAR" else "RESUELTO",
        "resolucion": resolucion,
        "resuelto_en": datetime.now().isoformat(),
        "resuelto_por_id": usuario.id if usuario and usuario.is_authenticated else None,
        "resuelto_por_nombre": usuario.nombre if usuario and usuario.is_authenticated else None,
        "requiere_revision": False,
    }
    if paciente:
        actualizada.update(
            {
                "paciente_id": paciente.id,
                "paciente_rut": paciente.rut,
                "paciente_nombre": paciente.nombre,
                "paciente_estado": paciente.estado,
                "paciente_id_ccr": paciente.id_ccr,
                "kine_asignado_nombre": paciente.kine_asignado.nombre if paciente.kine_asignado else None,
            }
        )
    return actualizada


ENCABEZADOS_PEGADO = [
    "FECHA DERIVACION",
    "NOMBRE",
    "RUT",
    "EDAD",
    "PERCAPITA / DESDE",
    "SECTOR OFICIAL",
    "SectorCesfam",
    "DIAGNOSTICO",
    "PROFESIONAL",
    "PRIORIDAD",
    "OBSERVACIONES",
    "KINE ASIGNADO",
    "ESTADO SUGERIDO",
    "CATEGORIA",
    "ASIGNADO HISTORICO",
]


def _filas_desde_texto_pegado(texto: str) -> list[list[str]]:
    lineas = [linea.rstrip("\r") for linea in str(texto or "").splitlines() if linea.strip()]
    if not lineas:
        raise ValueError("Debe pegar al menos una fila con encabezados.")

    muestra = lineas[:20]
    tabuladores = sum(linea.count("\t") for linea in muestra)
    punto_y_coma = sum(linea.count(";") for linea in muestra)
    comas = sum(linea.count(",") for linea in muestra)
    delimitador = "\t"
    if tabuladores == 0:
        delimitador = ";" if punto_y_coma >= comas else ","

    filas = [
        [str(celda).strip() for celda in fila]
        for fila in csv.reader(lineas, delimiter=delimitador)
    ]
    filas = [fila for fila in filas if any(celda.strip() for celda in fila)]
    filas = [fila for fila in filas if not _fila_basura_pegado(fila)]
    inicio_util = next(
        (
            index
            for index, fila in enumerate(filas)
            if _fila_contiene_encabezado_pegado(fila) or _fila_parece_dato_pegado(fila)
        ),
        None,
    )
    if inicio_util is not None:
        filas = filas[inicio_util:]

    if filas and not _filas_tienen_encabezado_pegado(filas) and _fila_parece_dato_pegado(filas[0]):
        filas_normalizadas = [_normalizar_fila_datos_pegada(fila) for fila in filas]
        filas = [fila.copy() for fila in ENCABEZADO_DERIVACIONES_PEGADO] + filas_normalizadas

    if len(filas) < 2:
        raise ValueError("Debe pegar encabezados y al menos un registro.")
    return filas


def _archivo_desde_texto_pegado(texto: str):
    filas = _filas_desde_texto_pegado(texto)
    wb = Workbook()
    ws = wb.active
    ws.title = "IMPORTAR"
    for fila in filas:
        ws.append(fila)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return ContentFile(buffer.getvalue(), name="pegado_planilla_ccr.xlsx")


def _agregar_diagnosticos_recurrentes(registros: list[dict], usuario) -> None:
    movimientos: list[MovimientoPaciente] = []
    pacientes_actualizados: list[Paciente] = []
    vistos: set[tuple[int, str]] = set()

    for registro in registros:
        if not registro.get("es_duplicado") or not registro.get("rut"):
            continue
        diagnostico = str(registro.get("diagnostico") or "").strip()
        if not diagnostico:
            continue
        paciente = _paciente_vinculado_por_registro(registro)
        if not paciente:
            continue
        key = (paciente.id, diagnostico.upper())
        if key in vistos:
            continue
        vistos.add(key)

        diagnostico_actual = paciente.diagnostico or ""
        if diagnostico.upper() not in diagnostico_actual.upper():
            paciente.diagnostico = (
                f"{diagnostico_actual}\n{diagnostico}"
                if diagnostico_actual.strip()
                else diagnostico
            )
            pacientes_actualizados.append(paciente)

        movimientos.append(
            MovimientoPaciente(
                paciente=paciente,
                usuario=usuario,
                estado_anterior=None,
                estado_nuevo=paciente.estado,
                notas=(
                    "Diagnóstico agregado desde importación recurrente: "
                    f"{diagnostico}. Hoja {registro.get('hoja') or '-'}, fila {registro.get('fila') or '-'}."
                ),
            )
        )

    if pacientes_actualizados:
        Paciente.objects.bulk_update(pacientes_actualizados, ["diagnostico"], batch_size=200)
    if movimientos:
        MovimientoPaciente.objects.bulk_create(movimientos, batch_size=200)


def _procesar_importacion_archivo(
    *,
    request,
    archivo,
    mes_solicitado: int | None,
    anio_fallback: int,
    forzar_reemplazo: bool,
    modo_suplementar: bool,
):
    hoy = date.today()
    resultado = parsear_derivaciones(
        archivo,
        mes_objetivo=mes_solicitado,
        anio_objetivo=anio_fallback if mes_solicitado else None,
    )
    periodos_detectados = _periodos_desde_resultado(
        resultado,
        anio_fallback=anio_fallback,
        mes_fallback=mes_solicitado,
    )

    conflictos = []
    importaciones_previas: dict[tuple[int, int], ImportacionMensual] = {}
    for (mes_num, anio_periodo), count in periodos_detectados.items():
        if count <= 0:
            continue
        importacion_previa = (
            ImportacionMensual.objects.filter(
                _periodo_q(mes_num, anio_periodo),
                estado__in=[
                    ImportacionMensual.Estado.COMPLETADO,
                    ImportacionMensual.Estado.CON_ERRORES,
                ],
            )
            .order_by("-fecha_subida")
            .first()
        )
        if importacion_previa:
            importaciones_previas[(mes_num, anio_periodo)] = importacion_previa
            conflictos.append(
                {
                    "hoja": f"{MESES_LABEL.get(mes_num, mes_num)} {anio_periodo}",
                    "mes": mes_num,
                    "anio": anio_periodo,
                    "importados_previos": importacion_previa.registros_importados,
                    "fecha_subida_previa": importacion_previa.fecha_subida.isoformat(),
                    "importacion_id": importacion_previa.id,
                }
            )

    if conflictos and not forzar_reemplazo and not modo_suplementar:
        return Response(
            {
                "tipo": "conflicto_mes",
                "mensaje": "Ya existen datos importados para estos meses.",
                "conflictos": conflictos,
                "pregunta": "¿Desea reemplazar los registros existentes de estos meses?",
            },
            status=status.HTTP_409_CONFLICT,
        )

    with transaction.atomic():
        if forzar_reemplazo and conflictos:
            _vaciar_importaciones(list(importaciones_previas.values()), borrar_importacion=False)
            archivo.seek(0)
            resultado = parsear_derivaciones(
                archivo,
                mes_objetivo=mes_solicitado,
                anio_objetivo=anio_fallback if mes_solicitado else None,
            )
            periodos_detectados = _periodos_desde_resultado(
                resultado,
                anio_fallback=anio_fallback,
                mes_fallback=mes_solicitado,
            )

        archivo.seek(0)
        archivo_bytes = archivo.read()
        archivo_nombre = getattr(archivo, "name", "") or "importacion_ccr.xlsx"
        nuevas_importaciones: dict[tuple[int, int], ImportacionMensual] = {}
        registros_por_periodo = _agrupar_registros_por_periodo(
            resultado.get("registros", []),
            periodos=periodos_detectados,
            mes_fallback=mes_solicitado,
            anio_fallback=anio_fallback,
        )
        errores_por_periodo = _agrupar_registros_por_periodo(
            resultado.get("errores", []),
            periodos=periodos_detectados,
            mes_fallback=mes_solicitado,
            anio_fallback=anio_fallback,
        )

        for (mes_num, anio_periodo), count in periodos_detectados.items():
            if count <= 0:
                continue
            registros_periodo = registros_por_periodo.get((mes_num, anio_periodo), [])
            duplicados_periodo = sum(
                1 for registro in registros_periodo if registro.get("estado") == "DUPLICADO"
            )
            errores_periodo = errores_por_periodo.get((mes_num, anio_periodo), [])

            nueva_importacion = ImportacionMensual.objects.create(
                archivo=ContentFile(archivo_bytes, name=archivo_nombre),
                archivo_nombre=archivo_nombre,
                mes=hoy.month,
                anio=hoy.year,
                mes_datos=mes_num,
                anio_datos=anio_periodo,
                usuario=request.user,
                estado=(
                    ImportacionMensual.Estado.COMPLETADO
                    if not errores_periodo
                    else ImportacionMensual.Estado.CON_ERRORES
                ),
                total_registros=count,
                registros_importados=resultado["importados"],
                duplicados=duplicados_periodo,
                errores=errores_periodo,
            )
            nuevas_importaciones[(mes_num, anio_periodo)] = nueva_importacion

        if not nuevas_importaciones and resultado["total"] > 0 and mes_solicitado:
            nueva_importacion = ImportacionMensual.objects.create(
                archivo=ContentFile(archivo_bytes, name=archivo_nombre),
                archivo_nombre=archivo_nombre,
                mes=hoy.month,
                anio=hoy.year,
                mes_datos=mes_solicitado,
                anio_datos=anio_fallback,
                usuario=request.user,
                estado=ImportacionMensual.Estado.CON_ERRORES,
                total_registros=resultado["total"],
                registros_importados=0,
                duplicados=resultado["duplicados"],
                errores=resultado["errores"],
            )
            nuevas_importaciones[(mes_solicitado, anio_fallback)] = nueva_importacion

        if forzar_reemplazo and conflictos:
            for periodo, importacion_previa in importaciones_previas.items():
                importacion_previa.estado = ImportacionMensual.Estado.REEMPLAZADO
                importacion_previa.reemplazada_por = nuevas_importaciones.get(periodo)
                importacion_previa.save(update_fields=["estado", "reemplazada_por"])

        pacientes_a_crear = resultado.get("pacientes", [])
        if pacientes_a_crear:
            import uuid as _uuid
            batch_prefix = _uuid.uuid4().hex[:6].upper()
            for i, p in enumerate(pacientes_a_crear):
                periodo_derivacion = (p.fecha_derivacion.month, p.fecha_derivacion.year)
                importacion_correspondiente = nuevas_importaciones.get(periodo_derivacion)
                if importacion_correspondiente:
                    p.importacion_origen = importacion_correspondiente
                elif nuevas_importaciones:
                    p.importacion_origen = list(nuevas_importaciones.values())[0]
                p.id_ccr = f"T{batch_prefix}{i + 1:05d}"

            creados = Paciente.objects.bulk_create(pacientes_a_crear, batch_size=200)
            for paciente in creados:
                paciente.id_ccr = _id_ccr_desde_pk(paciente.pk)
            Paciente.objects.bulk_update(creados, ["id_ccr"], batch_size=200)

            movimientos_iniciales = []
            for paciente in creados:
                estado_sugerido = getattr(
                    paciente,
                    "_importacion_estado_sugerido",
                    paciente.estado,
                )
                if estado_sugerido == Paciente.Estado.PENDIENTE:
                    continue
                responsable = getattr(paciente, "_importacion_kine_detectado", "") or "Sin responsable detectado"
                movimientos_iniciales.append(
                    MovimientoPaciente(
                        paciente=paciente,
                        usuario=request.user,
                        estado_anterior=None,
                        estado_nuevo=paciente.estado,
                        notas=(
                            "Carga inicial desde planilla CCR. "
                            "Paciente incluido en preorden histórico y/o asignación previa. "
                            f"Estado sugerido: {estado_sugerido}. "
                            f"Responsable detectado: {responsable}."
                        ),
                    )
                )
            if movimientos_iniciales:
                MovimientoPaciente.objects.bulk_create(movimientos_iniciales, batch_size=200)
            resultado["importados"] = len(creados)

        _agregar_diagnosticos_recurrentes(resultado.get("registros", []), request.user)

        ruts_duplicados = {
            reg["rut"]
            for reg in resultado.get("registros", [])
            if reg.get("es_duplicado") and reg.get("rut")
        }
        if ruts_duplicados:
            periodos_periodo = list(nuevas_importaciones.keys())
            exclude_periodos_q = Q()
            for mes_periodo, anio_periodo in periodos_periodo:
                exclude_periodos_q |= Q(
                    importacion_origen__mes_datos=mes_periodo,
                    importacion_origen__anio_datos=anio_periodo,
                )

            pats_a_actualizar = Paciente.objects.filter(
                rut__in=ruts_duplicados,
                estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE],
            )
            if exclude_periodos_q:
                pats_a_actualizar = pats_a_actualizar.exclude(exclude_periodos_q)

            pats_data = list(pats_a_actualizar.values_list("id", "estado"))
            if pats_data:
                pats_a_actualizar.update(n_meses_espera=F("n_meses_espera") + 1)
                meses_str = ", ".join(
                    f"{MESES_LABEL.get(m, str(m))} {anio}"
                    for m, anio in periodos_periodo
                )
                movimientos = [
                    MovimientoPaciente(
                        paciente_id=pid,
                        usuario=request.user,
                        estado_anterior=None,
                        estado_nuevo=estado,
                        notas=f"Registrado nuevamente en lista de espera: {meses_str}",
                    )
                    for pid, estado in pats_data
                ]
                MovimientoPaciente.objects.bulk_create(movimientos, batch_size=200)

        _registrar_observaciones_revision(
            registros=resultado.get("registros", []),
            importaciones=nuevas_importaciones,
            mes_fallback=mes_solicitado,
            anio_fallback=anio_fallback,
            usuario=request.user,
        )

        for importacion in nuevas_importaciones.values():
            importacion.registros_importados = importacion.pacientes_creados.count()
            importacion.save(update_fields=["registros_importados"])

        resultado.pop("pacientes", None)
        resultado.pop("registros", None)
        return Response(resultado, status=status.HTTP_201_CREATED)


class ImportarDerivacionesView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        serializer = ImportacionDerivacionesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        hoy = date.today()
        return _procesar_importacion_archivo(
            request=request,
            archivo=serializer.validated_data["archivo"],
            mes_solicitado=serializer.validated_data.get("mes"),
            anio_fallback=serializer.validated_data.get("anio", hoy.year),
            forzar_reemplazo=_bool_from_request(request.data.get("forzar_reemplazo", False)),
            modo_suplementar=_bool_from_request(request.data.get("modo_suplementar", False)),
        )


class PrevisualizarDerivacionesPegadasView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        try:
            archivo = _archivo_desde_texto_pegado(request.data.get("texto", ""))
            mes_solicitado = request.data.get("mes")
            anio_solicitado = request.data.get("anio")
            try:
                mes_solicitado = int(mes_solicitado) if mes_solicitado else None
                anio_solicitado = int(anio_solicitado) if anio_solicitado else None
            except (TypeError, ValueError):
                return Response({"detail": "Mes o año inválido."}, status=status.HTTP_400_BAD_REQUEST)
            resultado = previsualizar_derivaciones(
                archivo,
                mes_objetivo=mes_solicitado,
                anio_objetivo=anio_solicitado if mes_solicitado else None,
            )
            return Response(resultado, status=status.HTTP_200_OK)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class ImportarDerivacionesPegadasView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        try:
            archivo = _archivo_desde_texto_pegado(request.data.get("texto", ""))
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        hoy = date.today()
        mes_solicitado = request.data.get("mes")
        anio_fallback = request.data.get("anio") or hoy.year
        try:
            mes_solicitado = int(mes_solicitado) if mes_solicitado else None
            anio_fallback = int(anio_fallback)
        except (TypeError, ValueError):
            return Response({"detail": "Mes o año inválido."}, status=status.HTTP_400_BAD_REQUEST)

        return _procesar_importacion_archivo(
            request=request,
            archivo=archivo,
            mes_solicitado=mes_solicitado,
            anio_fallback=anio_fallback,
            forzar_reemplazo=_bool_from_request(request.data.get("forzar_reemplazo", False)),
            modo_suplementar=_bool_from_request(request.data.get("modo_suplementar", False)),
        )


class PrevisualizarDerivacionesView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        serializer = ImportacionDerivacionesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        archivo = serializer.validated_data["archivo"]
        mes_solicitado = serializer.validated_data.get("mes")
        anio_solicitado = serializer.validated_data.get("anio")
        resultado = previsualizar_derivaciones(
            archivo,
            mes_objetivo=mes_solicitado,
            anio_objetivo=anio_solicitado if mes_solicitado else None,
        )
        return Response(resultado, status=status.HTTP_200_OK)


class HistorialImportacionesView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request):
        historial = (
            ImportacionMensual.objects.select_related("usuario", "reemplazada_por")
            .all()
        )
        return Response([_serialize_importacion(item) for item in historial])


class ObservacionesRevisionImportacionView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request):
        tipo = request.query_params.get("tipo")
        estado_revision = request.query_params.get("estado", "PENDIENTE")
        mes = request.query_params.get("mes")
        anio = request.query_params.get("anio")

        importaciones = (
            ImportacionMensual.objects.select_related("usuario")
            .exclude(estado=ImportacionMensual.Estado.REEMPLAZADO)
            .order_by("-fecha_subida")
        )
        if mes and anio:
            try:
                importaciones = importaciones.filter(_periodo_q(int(mes), int(anio)))
            except ValueError:
                return Response(
                    {"detail": "Mes o año inválido."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        items: list[dict] = []
        total_pendientes = 0
        total_resueltos = 0
        total_descartados = 0
        for importacion in importaciones:
            periodo_mes, periodo_anio = _mes_y_anio_referencia(importacion)
            observaciones = _observaciones_revision_persistidas(importacion)

            for index, observacion in enumerate(observaciones):
                estado_item = observacion.get("estado_revision") or "PENDIENTE"
                if estado_item == "PENDIENTE":
                    total_pendientes += 1
                elif estado_item == "RESUELTO":
                    total_resueltos += 1
                elif estado_item == "DESCARTADO":
                    total_descartados += 1
                if tipo and observacion.get("tipo") != tipo:
                    continue
                if estado_revision != "TODOS" and estado_item != estado_revision:
                    continue
                items.append(
                    {
                        "id": f"{importacion.id}-{index}",
                        "importacion_id": importacion.id,
                        "revision_index": index,
                        "fecha_subida": importacion.fecha_subida.isoformat(),
                        "usuario_nombre": importacion.usuario.nombre if importacion.usuario else None,
                        "mes": periodo_mes,
                        "anio": periodo_anio,
                        "mes_label": MESES_LABEL.get(periodo_mes, str(periodo_mes)),
                        **observacion,
                    }
                )

        return Response(
            {
                "total": len(items),
                "pendientes": total_pendientes,
                "resueltos": total_resueltos,
                "descartados": total_descartados,
                "items": items,
            }
        )


class ObservacionRevisionDetalleView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    @transaction.atomic
    def patch(self, request, importacion_id: int, index: int):
        try:
            index_int = int(index)
        except (TypeError, ValueError):
            return Response(
                {"detail": "Índice de observación inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            importacion = ImportacionMensual.objects.get(pk=importacion_id)
        except ImportacionMensual.DoesNotExist:
            return Response(
                {"detail": "La importación no existe."},
                status=status.HTTP_404_NOT_FOUND,
            )

        resolved = _resolve_observacion(importacion, index_int)
        if not resolved:
            return Response(
                {"detail": "La observación no existe."},
                status=status.HTTP_404_NOT_FOUND,
            )

        observaciones, observacion = resolved
        accion = str(request.data.get("accion") or "").upper().strip()
        resolucion = str(request.data.get("resolucion") or "").strip()
        paciente = None

        if accion not in {"DESCARTAR", "COMPLETAR"}:
            return Response(
                {"detail": "Acción inválida."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if accion == "DESCARTAR":
            if not resolucion:
                resolucion = "Descartado por funcionario."
            observaciones[index_int] = _actualizar_observacion_resuelta(
                observacion=observacion,
                accion=accion,
                usuario=request.user,
                resolucion=resolucion,
            )
            importacion.observaciones_revision = observaciones
            importacion.save(update_fields=["observaciones_revision"])
            return Response({"item": observaciones[index_int]}, status=status.HTTP_200_OK)

        if accion == "COMPLETAR":
            paciente_data = request.data.get("paciente") or {}
            nombre = str(paciente_data.get("nombre") or observacion.get("nombre") or "").strip()
            rut = _normalizar_rut(paciente_data.get("rut") or observacion.get("rut"))
            diagnostico = str(paciente_data.get("diagnostico") or observacion.get("diagnostico") or "").strip()
            fecha_raw = str(
                paciente_data.get("fecha_derivacion")
                or observacion.get("fecha_derivacion")
                or observacion.get("fecha_original")
                or ""
            ).strip()
            profesional = str(paciente_data.get("profesional") or observacion.get("profesional") or "No informado").strip()
            percapita_desde = str(paciente_data.get("percapita_desde") or observacion.get("percapita_desde") or "").strip()
            sector_oficial = str(paciente_data.get("sector_oficial") or observacion.get("sector_oficial") or "").strip().upper()
            sector_cesfam = str(paciente_data.get("sector_cesfam") or observacion.get("sector_cesfam") or "").strip().upper()
            asignado_historico = bool(observacion.get("asignado_historico"))
            prioridad = prioridad_normalizada(str(paciente_data.get("prioridad") or observacion.get("prioridad") or "MODERADA"))
            categoria_raw = str(paciente_data.get("categoria") or observacion.get("categoria") or "").strip().upper()
            observaciones_texto = str(paciente_data.get("observaciones") or observacion.get("observaciones") or "").strip()
            try:
                edad = int(float(paciente_data.get("edad") or observacion.get("edad") or 0))
                fecha_derivacion = _parse_fecha_revisada(fecha_raw)
            except (TypeError, ValueError) as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

            if not nombre or not rut or not diagnostico:
                return Response(
                    {"detail": "Debe completar nombre, RUT y diagnóstico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            paciente = (
                Paciente.objects.filter(
                    rut=rut,
                    fecha_derivacion=fecha_derivacion,
                    diagnostico__iexact=diagnostico,
                )
                .order_by("-id")
                .first()
            )

            if not paciente:
                paciente = Paciente.objects.create(
                    fecha_derivacion=fecha_derivacion,
                    percapita_desde=percapita_desde,
                    sector_oficial=sector_oficial,
                    sector_cesfam=sector_cesfam,
                    asignado_historico=asignado_historico,
                    nombre=nombre.upper(),
                    rut=rut,
                    edad=edad,
                    diagnostico=diagnostico,
                    profesional=profesional,
                    prioridad=prioridad,
                    categoria=categoria_raw if categoria_raw in Paciente.Categoria.values else categoria_por_diagnostico(diagnostico, edad),
                    observaciones=observaciones_texto,
                    importacion_origen=importacion,
                )
                movimiento_nota = "Ficha operativa creada desde revisión de importación."
            else:
                movimiento_nota = "Observación vinculada a ficha operativa existente desde revisión de importación."

            observacion = {
                **observacion,
                "nombre": nombre.upper(),
                "rut": rut,
                "fecha_derivacion": fecha_derivacion.strftime("%d/%m/%Y"),
                "edad": edad,
                "diagnostico": diagnostico,
                "prioridad": prioridad,
                "percapita_desde": percapita_desde,
                "sector_oficial": sector_oficial,
                "sector_cesfam": sector_cesfam,
                "asignado_historico": asignado_historico,
                "profesional": profesional,
                "categoria": categoria_raw if categoria_raw in Paciente.Categoria.values else categoria_por_diagnostico(diagnostico, edad),
                "observaciones": observaciones_texto,
            }
            if not resolucion:
                resolucion = "Datos completados y ficha operativa vinculada."

            MovimientoPaciente.objects.create(
                paciente=paciente,
                usuario=request.user,
                estado_anterior=None,
                estado_nuevo=paciente.estado,
                notas=f"{movimiento_nota} Corte {observacion.get('periodo_label')}. Resolución: {resolucion}",
            )

        observaciones[index_int] = _actualizar_observacion_resuelta(
            observacion=observacion,
            accion=accion,
            usuario=request.user,
            resolucion=resolucion,
            paciente=paciente,
        )
        importacion.observaciones_revision = observaciones
        importacion.save(update_fields=["observaciones_revision"])
        return Response({"item": observaciones[index_int]}, status=status.HTTP_200_OK)


class ExportarHistorialImportacionesMesView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request, mes: int, anio: int):
        mes = int(mes)
        anio = int(anio)
        periodo = f"{MESES_LABEL.get(mes, mes)} {anio}"
        pacientes = (
            Paciente.objects.select_related("kine_asignado", "importacion_origen")
            .filter(fecha_derivacion__month=mes, fecha_derivacion__year=anio)
            .order_by("fecha_derivacion", "nombre")
        )
        workbook = crear_excel_pacientes(
            pacientes,
            titulo=f"Corte mensual CCR - {periodo}",
            subtitulo="Pacientes derivados en el periodo",
            filtros={"mes": str(mes), "anio": str(anio)},
            incluir_importacion=True,
            periodo=periodo,
            mensaje_vacio="Sin pacientes para este corte.",
        )
        nombre_mes = MESES_LABEL.get(mes, str(mes)).lower()
        return excel_response(workbook, f"corte-ccr-{nombre_mes}-{anio}.xlsx")


class HistorialImportacionesMesView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request, mes: int, anio: int):
        mes = int(mes)
        anio = int(anio)
        historial = (
            ImportacionMensual.objects.select_related("usuario", "reemplazada_por")
            .filter(_periodo_q(mes, anio))
            .order_by("-fecha_subida")
        )
        return Response(
            {
                "mes": mes,
                "anio": anio,
                "mes_label": MESES_LABEL.get(mes, str(mes)),
                "periodo_label": f"{MESES_LABEL.get(mes, mes)} {anio}",
                "pacientes_actuales_del_corte": _resumen_pacientes_corte(mes, anio),
                "items": [_serialize_importacion(item) for item in historial],
            }
        )

    @transaction.atomic
    def delete(self, request, mes: int, anio: int):
        try:
            mes = int(mes)
            anio = int(anio)
        except (TypeError, ValueError):
            return Response(
                {"detail": "Periodo inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if mes < 1 or mes > 12:
            return Response(
                {"detail": "Mes fuera de rango."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        historial_qs = ImportacionMensual.objects.filter(_periodo_q(mes, anio)).order_by(
            "-fecha_subida"
        )
        if not historial_qs.exists():
            return Response(
                {"detail": "No hay importaciones para ese periodo."},
                status=status.HTTP_404_NOT_FOUND,
            )

        importaciones = list(historial_qs)
        resumen = _vaciar_importaciones(importaciones, borrar_importacion=True)

        return Response(
            {
                "mes": mes,
                "anio": anio,
                **resumen,
            },
            status=status.HTTP_200_OK,
        )


class ImportacionMensualDetalleView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    @transaction.atomic
    def delete(self, request, pk: int):
        try:
            importacion = ImportacionMensual.objects.get(pk=pk)
        except ImportacionMensual.DoesNotExist:
            return Response(
                {"detail": "El corte no existe."},
                status=status.HTTP_404_NOT_FOUND,
            )

        mes, anio = _mes_y_anio_referencia(importacion)
        resumen = _vaciar_importaciones([importacion], borrar_importacion=True)

        return Response(
            {
                "id": int(pk),
                "mes": mes,
                "anio": anio,
                **resumen,
            },
            status=status.HTTP_200_OK,
        )


class PlantillaImportacionView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "IMPORTAR"

        ws.merge_cells("A1:N1")
        ws["A1"] = "Derivación a Centro Comunitario de Rehabilitación Cesfam Dr. Alberto Reyes"
        ws["A1"].font = Font(bold=True, size=11)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers_f2 = {
            "A2": "FECHA DERIV",
            "B2": "SECTOR OFICIAL",
            "C2": "SectorCesfam",
            "D2": "NOMBRE",
            "E2": "RUT",
            "F2": "EDAD",
            "G2": "DIAGNÓSTICO",
            "H2": "PROFESIONAL",
            "I2": "PRIORIDAD",
            "J2": "OBSERVACIONES",
            "K2": "KINE ASIGNADO",
            "L2": "ESTADO SUGERIDO",
            "M2": "CATEGORIA",
            "N2": "ASIGNADO HISTORICO",
        }
        fill_header = PatternFill("solid", fgColor="1B5E3B")
        font_header = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell_ref, val in headers_f2.items():
            c = ws[cell_ref]
            c.value = val
            c.font = font_header
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[2].height = 24

        ejemplos = {
            "A3": "DD/MM/YYYY",
            "B3": "AZUL / ROJO / VERDE",
            "C3": "AZUL / CECOSF EL SANTO",
            "D3": "NOMBRE COMPLETO",
            "E3": "12345678-9",
            "F3": "65",
            "G3": "LUMBAGO",
            "H3": "KINESIOLOGO",
            "I3": "ALTA / MEDIANA / MODERADA",
            "J3": "Observaciones opcionales",
            "K3": "SEBA C / MANE / PILAR...",
            "L3": "PENDIENTE / INGRESADO",
            "M3": "LUMBAGOS / MAS65",
            "N3": "SI / NO",
        }
        fill_ejemplo = PatternFill("solid", fgColor="E8F5EE")
        font_ejemplo = Font(italic=True, color="3A5A3A", size=9)
        for cell_ref, val in ejemplos.items():
            c = ws[cell_ref]
            c.value = val
            c.font = font_ejemplo
            c.fill = fill_ejemplo
            c.alignment = Alignment(horizontal="center")
            c.border = border

        datos_ejemplo = [
            (
                "15/01/2025",
                "AZUL",
                "AZUL",
                "JUAN PÉREZ GONZÁLEZ",
                "12345678-9",
                65,
                "LUMBAGO",
                "KINESIOLOGO",
                "ALTA",
                "Dolor crónico lumbar",
                "SEBA C",
                "PENDIENTE",
                "LUMBAGOS",
                "NO",
            ),
            (
                "20/01/2025",
                "CECOSF EL SANTO",
                "CECOSF EL SANTO",
                "MARÍA SOTO RAMÍREZ",
                "9876543-2",
                72,
                "GONARTROSIS (GES)",
                "KINESIOLOGO",
                "MEDIANA",
                "",
                "MANE",
                "INGRESADO",
                "MAS65",
                "SI",
            ),
            (
                "22/01/2025",
                "ROJO",
                "ROJO",
                "PEDRO MUÑOZ VEGA",
                "15432198-K",
                45,
                "HOMBRO DOLOROSO",
                "KINESIOLOGO",
                "MODERADA",
                "ECO compatible",
                "",
                "",
                "HOMBROS",
                "NO",
            ),
        ]
        for i, fila in enumerate(datos_ejemplo, start=4):
            for j, val in enumerate(fila, start=1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = Font(size=10)
                c.border = border
                if j == 1:
                    c.number_format = "DD/MM/YYYY"

        anchos = [14, 20, 20, 30, 14, 7, 28, 18, 16, 40, 20, 20, 28, 18]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[chr(64 + i)].width = ancho

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Plantilla_Derivaciones_CCR.xlsx"'
        return response


class ResetPoblacionView(APIView):
    """
    DELETE /api/importar/reset
    Elimina TODOS los pacientes sin responsable CCR asignado.
    Los pacientes ASIGNADOS se conservan intactos.
    Tambien limpia el historial de importaciones.
    Solo accesible para ADMIN.
    """
    permission_classes = [IsAdminOrAdministrativoRole]

    @transaction.atomic
    def delete(self, request):
        if not (request.user.is_authenticated and getattr(request.user, 'rol', None) == 'ADMIN'):
            return Response(
                {'detail': 'Solo el administrador puede resetear la poblacion.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Solo eliminamos pacientes que:
        # 1. No tengan responsable CCR asignado.
        # 2. Estén en estado PENDIENTE o RESCATE
        # (Los INGRESADOS, EGRESADOS, ALTA, DERIVADO, ABANDONO se mantienen)
        sin_asignar_qs = Paciente.objects.filter(
            kine_asignado__isnull=True,
            estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE]
        )
        total_eliminados = sin_asignar_qs.count()
        sin_asignar_qs.delete()

        importaciones_qs = ImportacionMensual.objects.all()
        total_importaciones = importaciones_qs.count()
        for imp in importaciones_qs:
            if imp.archivo:
                try:
                    imp.archivo.delete(save=False)
                except Exception:
                    pass
        importaciones_qs.delete()

        return Response(
            {
                'pacientes_eliminados': total_eliminados,
                'importaciones_eliminadas': total_importaciones,
                'mensaje': f'Se eliminaron {total_eliminados} pacientes sin asignar y {total_importaciones} registros de importacion.',
            },
            status=status.HTTP_200_OK,
        )
