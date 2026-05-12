from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.importar.parser import previsualizar_derivaciones


IMPORTAR_HEADERS = [
    "FECHA DERIV",
    "SECTOR OFICIAL",
    "SectorCesfam",
    "NOMBRE",
    "RUT",
    "EDAD",
    "DIAGNÓSTICO",
    "PROFESIONAL",
    "PRIORIDAD",
    "OBSERVACIONES",
    "KINE ASIGNADO",
    "ESTADO SUGERIDO",
    "CATEGORIA",
    "ASIGNADO HISTORICO",
]

REVISION_HEADERS = [
    "MOTIVO REVISION",
    "FECHA DERIV",
    "SECTOR OFICIAL",
    "SectorCesfam",
    "NOMBRE",
    "RUT",
    "EDAD",
    "DIAGNÓSTICO",
    "PROFESIONAL",
    "PRIORIDAD",
    "OBSERVACIONES",
    "KINE DETECTADO",
    "ESTADO SUGERIDO",
    "CATEGORIA",
    "ASIGNADO HISTORICO",
    "HOJA ORIGEN",
    "FILA ORIGEN",
]


def _valor(registro: dict, key: str, default=""):
    value = registro.get(key, default)
    return default if value is None else value


def _fila_importar(registro: dict) -> list:
    return [
        _valor(registro, "fecha_derivacion"),
        _valor(registro, "sector_oficial"),
        _valor(registro, "sector_cesfam"),
        _valor(registro, "nombre"),
        _valor(registro, "rut"),
        _valor(registro, "edad", 0),
        _valor(registro, "diagnostico"),
        _valor(registro, "profesional"),
        _valor(registro, "prioridad"),
        _valor(registro, "observaciones"),
        _valor(registro, "kine_asignado"),
        _valor(registro, "estado_sugerido") or "PENDIENTE",
        _valor(registro, "categoria"),
        "SI" if registro.get("asignado_historico") else "NO",
    ]


def _motivo_revision(registro: dict) -> str:
    if registro.get("error"):
        return registro["error"]
    if registro.get("motivo_revision"):
        return registro["motivo_revision"]
    if registro.get("es_duplicado"):
        return "Paciente recurrente o duplicado en el sistema."
    return "Registro requiere revisión operativa."


def _fila_revision(registro: dict) -> list:
    return [
        _motivo_revision(registro),
        _valor(registro, "fecha_derivacion"),
        _valor(registro, "sector_oficial"),
        _valor(registro, "sector_cesfam"),
        _valor(registro, "nombre"),
        _valor(registro, "rut"),
        _valor(registro, "edad", 0),
        _valor(registro, "diagnostico"),
        _valor(registro, "profesional"),
        _valor(registro, "prioridad"),
        _valor(registro, "observaciones"),
        _valor(registro, "kine_asignado"),
        _valor(registro, "estado_sugerido") or "PENDIENTE",
        _valor(registro, "categoria"),
        "SI" if registro.get("asignado_historico") else "NO",
        _valor(registro, "hoja"),
        _valor(registro, "fila"),
    ]


def _aplicar_formato(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="EAF1FF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="173B8F")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    for index, header in enumerate(headers, start=1):
        ancho = min(max(len(header) + 4, 14), 34)
        ws.column_dimensions[get_column_letter(index)].width = ancho


def _escribir_hoja(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _aplicar_formato(ws, headers)


class Command(BaseCommand):
    help = "Limpia un Excel CCR y genera hojas IMPORTAR, REVISION y RESUMEN."

    def add_arguments(self, parser):
        parser.add_argument("--input", required=True, help="Ruta del Excel original.")
        parser.add_argument("--output", required=True, help="Ruta del Excel limpio de salida.")

    def handle(self, *args, **options):
        input_path = Path(options["input"]).expanduser()
        output_path = Path(options["output"]).expanduser()
        if not input_path.exists():
            raise CommandError(f"No existe el archivo de entrada: {input_path}")

        resultado = previsualizar_derivaciones(str(input_path))
        registros = resultado.get("registros", [])
        listos = [registro for registro in registros if registro.get("estado") == "OK"]
        revision = [registro for registro in registros if registro.get("estado") != "OK"]

        wb = Workbook()
        ws_importar = wb.active
        ws_importar.title = "IMPORTAR"
        _escribir_hoja(ws_importar, IMPORTAR_HEADERS, [_fila_importar(registro) for registro in listos])

        ws_revision = wb.create_sheet("REVISION")
        _escribir_hoja(ws_revision, REVISION_HEADERS, [_fila_revision(registro) for registro in revision])

        conteo_kine = Counter(_valor(registro, "kine_asignado") or "Sin kine" for registro in listos)
        conteo_sector_cesfam = Counter(_valor(registro, "sector_cesfam") or "Sin dato" for registro in listos)
        conteo_sector_oficial = Counter(_valor(registro, "sector_oficial") or "Sin dato" for registro in listos)
        conteo_categoria = Counter(_valor(registro, "categoria") or "Sin dato" for registro in listos)
        conteo_estado = Counter((_valor(registro, "estado_sugerido") or "PENDIENTE") for registro in listos)
        conteo_motivo = Counter(_motivo_revision(registro) for registro in revision)

        ws_resumen = wb.create_sheet("RESUMEN")
        ws_resumen.append(["Métrica", "Valor"])
        resumen_base = [
            ("Total registros leídos", resultado.get("total", 0)),
            ("Total listos para importar", len(listos)),
            ("Total enviados a revisión", len(revision)),
            ("Total con kine asignado", sum(1 for registro in listos if registro.get("kine_asignado"))),
            ("Total sin kine asignado", sum(1 for registro in listos if not registro.get("kine_asignado"))),
            ("Total asignado histórico", sum(1 for registro in listos if registro.get("asignado_historico"))),
        ]
        for item in resumen_base:
            ws_resumen.append(list(item))

        bloques = [
            ("Conteo por SectorCesfam", conteo_sector_cesfam),
            ("Conteo por SECTOR OFICIAL", conteo_sector_oficial),
            ("Conteo por CATEGORIA", conteo_categoria),
            ("Conteo por estado sugerido", conteo_estado),
            ("Conteo por kine asignado", conteo_kine),
            ("Conteo por motivo de revisión", conteo_motivo),
        ]
        for titulo, counter in bloques:
            ws_resumen.append([])
            ws_resumen.append([titulo, "Total"])
            for key, total in counter.most_common():
                ws_resumen.append([key, total])
        _aplicar_formato(ws_resumen, ["Métrica", "Valor"])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        self.stdout.write(self.style.SUCCESS(f"Excel limpio generado: {output_path}"))
        self.stdout.write(f"Listos para importar: {len(listos)}")
        self.stdout.write(f"En revisión: {len(revision)}")
