from datetime import date
from io import BytesIO, StringIO
import shutil
import tempfile

from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import override_settings
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from apps.importar.models import ImportacionMensual
from apps.pacientes.models import MovimientoPaciente, Paciente
from apps.usuarios.models import Usuario


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADERS_OFICIALES_IMPORTAR = [
    "FECHA DERIV",
    "SECTOR OFICIAL",
    "SectorCesfam",
    "NOMBRE",
    "RUT",
    "EDAD",
    "DIAGNÓSTICO",
    "PROFESIONAL",
    "PRIORIDAD",
    "OBSERVACIONES",
    "KINE ASIGNADO",
    "ESTADO SUGERIDO",
    "CATEGORIA",
    "ASIGNADO HISTORICO",
]


def excel_derivaciones(rows, sheet_name="JULIO"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(
        [
            "FECHA",
            "NOMBRE",
            "RUT",
            "EDAD",
            "DESDE",
            "DIAGNÓSTICO",
            "PROFESIONAL",
            "PRIORIDAD",
            "OBSERVACIONES",
        ]
    )
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def excel_derivaciones_con_headers(headers, rows, sheet_name="JULIO"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


class ImportacionesMensualesTests(APITestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override_settings = override_settings(MEDIA_ROOT=self.media_root)
        self.override_settings.enable()
        self.admin = Usuario.objects.create_user(
            rut="44444444-4",
            password="testpass",
            nombre="Admin Importaciones",
            rol=Usuario.Rol.ADMIN,
        )
        self.client.force_authenticate(self.admin)

    def tearDown(self):
        self.override_settings.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)
        super().tearDown()

    def archivo(self, rows, name="derivaciones.xlsx"):
        return SimpleUploadedFile(
            name,
            excel_derivaciones(rows),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def archivo_con_headers(self, headers, rows, name="derivaciones.xlsx"):
        return SimpleUploadedFile(
            name,
            excel_derivaciones_con_headers(headers, rows),
            content_type=EXCEL_CONTENT_TYPE,
        )

    def crear_kines_activos(self):
        datos = [
            ("101010101", "Sebastián Campos"),
            ("202020202", "Sebastián Salgado"),
            ("303030303", "Mane Sáez"),
            ("404040404", "Pilar Alarcón"),
            ("505050505", "Karen Torres"),
        ]
        for rut, nombre in datos:
            Usuario.objects.create_user(
                rut=rut,
                password="testpass",
                nombre=nombre,
                rol=Usuario.Rol.KINE,
                is_active=True,
            )

    def test_preview_no_crea_pacientes_y_entrega_resumen(self):
        archivo = self.archivo(
            [
                [date(2025, 7, 2), "Paciente Nuevo", "11111111-1", 50, "CESFAM", "Lumbago", "KINESIOLOGO", "ALTA", ""],
                ["sin fecha", "Paciente Error", "22222222-2", 60, "CESFAM", "Hombro", "KINESIOLOGO", "MODERADA", ""],
            ]
        )

        response = self.client.post(
            "/api/importar/previsualizar/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Paciente.objects.count(), 0)
        self.assertEqual(response.data["total"], 2)
        self.assertEqual(response.data["nuevos"], 1)
        self.assertEqual(response.data["errores_count"], 1)
        self.assertEqual(response.data["registros"][0]["tipo_revision"], "ADVERTENCIA")
        self.assertIn("Sector oficial vacío", response.data["registros"][0]["motivo_revision"])

    def test_preview_rechaza_archivo_que_no_es_xlsx(self):
        archivo = SimpleUploadedFile(
            "derivaciones.csv",
            b"fecha,nombre\n",
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/importar/previsualizar/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Paciente.objects.count(), 0)
        self.assertIn("archivo", response.data)

    def test_importacion_crea_pacientes_y_asocia_importacion(self):
        archivo = self.archivo(
            [
                [date(2025, 7, 2), "Paciente Uno", "11111111-1", 50, "CESFAM", "Lumbago", "KINESIOLOGO", "ALTA", ""],
                [date(2025, 7, 3), "Paciente Dos", "22222222-2", 55, "CESFAM", "Hombro", "KINESIOLOGO", "MODERADA", ""],
            ]
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["importados"], 2)
        importacion = ImportacionMensual.objects.get()
        self.assertEqual(importacion.registros_importados, 2)
        self.assertEqual(importacion.pacientes_creados.count(), 2)
        pacientes_importados = list(Paciente.objects.filter(importacion_origen=importacion))
        self.assertEqual(len(pacientes_importados), 2)
        for paciente in pacientes_importados:
            self.assertTrue(paciente.id_ccr)
            self.assertTrue(paciente.id_ccr.startswith("CCR-"))
            self.assertEqual(paciente.importacion_origen, importacion)

    def test_importa_columnas_nuevas_y_asigna_kines_activos(self):
        self.crear_kines_activos()
        archivo = self.archivo_con_headers(
            HEADERS_OFICIALES_IMPORTAR,
            [
                [date(2025, 7, 2), "AZUL", "SECTOR AZUL", "Paciente Campos", "12121212-1", 50, "Lumbago", "KINE", "ALTA", "", "SEBA C", "INGRESADO", "LUMBAGOS", "SI"],
                [date(2025, 7, 3), "ROJO", "CECOSF CERRO ESTANQUE", "Paciente Salgado", "13131313-1", 45, "Hombro", "KINE", "MODERADA", "", "SEBASTIÁN SALGADO", "PENDIENTE", "HOMBRO", "NO"],
                [date(2025, 7, 4), "VERDE", "EL SANTO", "Paciente Mane", "14141414-1", 65, "Rodilla", "KINE", "MEDIA", "", "MANE SÁEZ", "", "", "SI"],
            ],
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["importados"], 3)
        campos = Paciente.objects.get(rut="121212121")
        salgado = Paciente.objects.get(rut="131313131")
        mane = Paciente.objects.get(rut="141414141")
        self.assertEqual(campos.sector_oficial, "AZUL")
        self.assertEqual(campos.sector_cesfam, "AZUL")
        self.assertEqual(campos.categoria, Paciente.Categoria.LUMBAGOS)
        self.assertEqual(campos.kine_asignado.nombre, "Sebastián Campos")
        self.assertEqual(campos.estado, Paciente.Estado.INGRESADO)
        self.assertTrue(campos.asignado_historico)
        self.assertEqual(salgado.sector_cesfam, "CECOSF CERRO ESTANQUE")
        self.assertEqual(salgado.categoria, Paciente.Categoria.HOMBROS)
        self.assertEqual(salgado.kine_asignado.nombre, "Sebastián Salgado")
        self.assertFalse(salgado.asignado_historico)
        self.assertEqual(mane.sector_cesfam, "CECOSF EL SANTO")
        self.assertEqual(mane.kine_asignado.nombre, "Mane Sáez")
        self.assertTrue(mane.asignado_historico)

    def test_previsualizacion_formato_oficial_entrega_conteos_operativos(self):
        self.crear_kines_activos()
        archivo = self.archivo_con_headers(
            HEADERS_OFICIALES_IMPORTAR,
            [
                [date(2025, 7, 2), "AZUL", "AZUL", "Paciente Ingresado", "61616161-1", 50, "Lumbago", "KINE", "ALTA", "", "SEBA C", "INGRESADO", "LUMBAGOS", "SI"],
                [date(2025, 7, 3), "ROJO", "ROJO", "Paciente Pendiente", "62626262-1", 45, "Hombro", "KINE", "MODERADA", "", "", "PENDIENTE", "HOMBROS", "NO"],
            ],
        )

        response = self.client.post(
            "/api/importar/previsualizar/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total"], 2)
        self.assertEqual(response.data["validos"], 2)
        self.assertEqual(response.data["con_kine_asignado"], 1)
        self.assertEqual(response.data["sin_kine_asignado"], 1)
        self.assertEqual(response.data["ingresados"], 1)
        self.assertEqual(response.data["pendientes"], 1)
        self.assertEqual(response.data["asignado_historico"], 1)
        self.assertEqual(response.data["conteo_sector_cesfam"]["AZUL"], 1)
        self.assertEqual(response.data["conteo_categoria"]["LUMBAGOS"], 1)
        self.assertEqual(response.data["conteo_kine_asignado"]["Sebastián Campos"], 1)

    def test_kine_vacio_no_se_infiere_desde_observaciones(self):
        self.crear_kines_activos()
        archivo = self.archivo_con_headers(
            HEADERS_OFICIALES_IMPORTAR,
            [
                [date(2025, 7, 2), "AZUL", "AZUL", "Paciente Observacion", "63636363-1", 50, "Lumbago", "KINE", "ALTA", "Revisado por SEBA C", "", "PENDIENTE", "LUMBAGOS", "NO"],
            ],
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        paciente = Paciente.objects.get(rut="636363631")
        self.assertIsNone(paciente.kine_asignado)

    def test_kines_no_activos_quedan_sin_asignacion_y_en_revision(self):
        self.crear_kines_activos()
        headers = [
            "FECHA DERIV",
            "NOMBRE",
            "RUT",
            "EDAD",
            "DIAGNÓSTICO",
            "PROFESIONAL",
            "PRIORIDAD",
            "OBSERVACIONES",
            "KINE ASIGNADO",
        ]
        archivo = self.archivo_con_headers(
            headers,
            [
                [date(2025, 7, 2), "Paciente Benja", "15151515-1", 40, "Lumbago", "KINE", "MODERADA", "", "BENJA"],
                [date(2025, 7, 3), "Paciente Ignacia", "16161616-1", 41, "Hombro", "KINE", "MODERADA", "", "MARÍA IGNACIA"],
            ],
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["importados"], 2)
        self.assertEqual(Paciente.objects.filter(kine_asignado__isnull=True).count(), 2)
        importacion = ImportacionMensual.objects.get()
        self.assertEqual(len(importacion.observaciones_revision), 2)
        self.assertEqual(importacion.observaciones_revision[0]["tipo"], "ADVERTENCIA")
        self.assertIn("Kine no activo", importacion.observaciones_revision[0]["motivo"])

    def test_categoria_vacia_o_invalida_se_calcula_sin_bloquear(self):
        headers = [
            "FECHA DERIV",
            "SECTOR OFICIAL",
            "SectorCesfam",
            "NOMBRE",
            "RUT",
            "EDAD",
            "DIAGNÓSTICO",
            "PROFESIONAL",
            "PRIORIDAD",
            "OBSERVACIONES",
            "CATEGORIA",
        ]
        archivo = self.archivo_con_headers(
            headers,
            [
                [date(2025, 7, 2), "AZUL", "AZUL", "Paciente Sin Categoria", "31313131-1", 50, "Lumbago", "KINE", "ALTA", "", ""],
                [date(2025, 7, 3), "ROJO", "ROJO", "Paciente Categoria Mala", "32323232-1", 45, "Hombro", "KINE", "MODERADA", "", "NO EXISTE"],
            ],
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        sin_categoria = Paciente.objects.get(rut="313131311")
        categoria_mala = Paciente.objects.get(rut="323232321")
        self.assertEqual(sin_categoria.categoria, Paciente.Categoria.LUMBAGOS)
        self.assertEqual(categoria_mala.categoria, Paciente.Categoria.HOMBROS)
        importacion = ImportacionMensual.objects.get()
        self.assertIn("Categoría inválida", importacion.observaciones_revision[0]["motivo"])

    def test_estado_sugerido_ingresado_crea_movimiento_inicial(self):
        self.crear_kines_activos()
        headers = [
            "FECHA DERIV",
            "NOMBRE",
            "RUT",
            "EDAD",
            "DIAGNOSTICO",
            "PROFESIONAL",
            "PRIORIDAD",
            "OBSERVACIONES",
            "KINE ASIGNADO",
            "ESTADO SUGERIDO",
        ]
        archivo = self.archivo_con_headers(
            headers,
            [
                [date(2025, 7, 5), "Paciente Ingresado", "17171717-1", 50, "Lumbago", "KINE", "ALTA", "", "SEBA CAMPOS", "INGRESADO"],
                [date(2025, 7, 6), "Paciente Pendiente", "18181818-1", 50, "Lumbago", "KINE", "ALTA", "", "SEBA CAMPOS", ""],
            ],
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        ingresado = Paciente.objects.get(rut="171717171")
        pendiente = Paciente.objects.get(rut="181818181")
        self.assertEqual(ingresado.estado, Paciente.Estado.INGRESADO)
        self.assertEqual(pendiente.estado, Paciente.Estado.PENDIENTE)
        movimiento = MovimientoPaciente.objects.get(paciente=ingresado)
        self.assertEqual(movimiento.estado_anterior, None)
        self.assertEqual(movimiento.estado_nuevo, Paciente.Estado.INGRESADO)
        self.assertIn("Carga inicial desde planilla CCR", movimiento.notas)
        self.assertIn("preorden histórico", movimiento.notas)

    def test_limpia_excel_ccr_genera_importar_revision_y_resumen(self):
        self.crear_kines_activos()
        input_path = f"{self.media_root}/original.xlsx"
        output_path = f"{self.media_root}/CCR_precarga_limpia.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "JULIO"
        ws.append(HEADERS_OFICIALES_IMPORTAR)
        ws.append([date(2025, 7, 2), "AZUL", "AZUL", "Paciente Limpio", "19191919-1", 50, "Lumbago", "KINE", "ALTA", "", "SEBA C", "", "LUMBAGOS", "SI"])
        ws.append([date(2025, 7, 3), "ROJO", "ROJO", "Paciente Revision", "20202020-1", 50, "Lumbago", "KINE", "ALTA", "", "BENJA", "", "LUMBAGOS", "NO"])
        wb.save(input_path)

        call_command("limpiar_excel_ccr", input=input_path, output=output_path)

        salida = load_workbook(output_path)
        self.assertEqual(salida.sheetnames, ["IMPORTAR", "REVISION", "RESUMEN"])
        self.assertEqual(salida["IMPORTAR"].max_row, 2)
        self.assertEqual(salida["REVISION"].max_row, 2)
        headers_salida = [cell.value for cell in salida["IMPORTAR"][1]]
        self.assertEqual(headers_salida[:14], [
            "FECHA DERIV",
            "SECTOR OFICIAL",
            "SectorCesfam",
            "NOMBRE",
            "RUT",
            "EDAD",
            "DIAGNÓSTICO",
            "PROFESIONAL",
            "PRIORIDAD",
            "OBSERVACIONES",
            "KINE ASIGNADO",
            "ESTADO SUGERIDO",
            "CATEGORIA",
            "ASIGNADO HISTORICO",
        ])
        self.assertNotIn("PERCÁPITA / DESDE", headers_salida)
        self.assertNotIn("FUENTE ASIGNACION", headers_salida)
        self.assertNotIn("RECEPCION ORIGINAL", headers_salida)
        self.assertNotIn("SECCIÓN", headers_salida)
        self.assertEqual(salida["IMPORTAR"]["B2"].value, "AZUL")
        self.assertEqual(salida["IMPORTAR"]["C2"].value, "AZUL")
        self.assertEqual(salida["IMPORTAR"]["N2"].value, "SI")
        self.assertIn("Kine no activo", salida["REVISION"]["A2"].value)

    def test_recurrentes_no_crean_duplicados_y_quedan_en_revision(self):
        Paciente.objects.create(
            id_ccr="EXISTENTE-1",
            fecha_derivacion=date(2025, 7, 2),
            percapita_desde="CESFAM",
            nombre="PACIENTE EXISTENTE",
            rut="111111111",
            edad=50,
            diagnostico="Lumbago",
            profesional="KINESIOLOGO",
            prioridad=Paciente.Prioridad.ALTA,
            categoria=Paciente.Categoria.LUMBAGOS,
        )
        archivo = self.archivo(
            [
                [date(2025, 7, 2), "Paciente Existente", "11111111-1", 50, "CESFAM", "Lumbago", "KINESIOLOGO", "ALTA", ""],
            ]
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["importados"], 0)
        self.assertEqual(response.data["recurrentes"], 1)
        self.assertEqual(Paciente.objects.count(), 1)
        importacion = ImportacionMensual.objects.get()
        self.assertEqual(importacion.duplicados, 1)
        self.assertEqual(importacion.observaciones_revision[0]["tipo"], "RECURRENTE")

    def test_errores_quedan_guardados_en_importacion(self):
        archivo = self.archivo(
            [
                ["fecha mala", "Paciente Error", "33333333-3", 45, "CESFAM", "Rodilla", "KINESIOLOGO", "MODERADA", ""],
            ]
        )

        response = self.client.post(
            "/api/importar/derivaciones/",
            {"archivo": archivo, "mes": 7, "anio": 2025},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        importacion = ImportacionMensual.objects.get()
        self.assertEqual(importacion.estado, ImportacionMensual.Estado.CON_ERRORES)
        self.assertEqual(len(importacion.errores), 1)
        self.assertEqual(importacion.observaciones_revision[0]["tipo"], "ERROR")

    def test_revision_get_y_patch_actualiza_estado(self):
        importacion = ImportacionMensual.objects.create(
            archivo=ContentFile(b"test", name="revision.xlsx"),
            archivo_nombre="revision.xlsx",
            mes=5,
            anio=2026,
            mes_datos=7,
            anio_datos=2025,
            usuario=self.admin,
            observaciones_revision=[
                {
                    "tipo": "ERROR",
                    "tipo_label": "Error de datos",
                    "periodo_label": "Julio 2025",
                    "archivo_nombre": "revision.xlsx",
                    "hoja": "JULIO",
                    "fila": 2,
                    "motivo": "Fecha inválida",
                    "accion": "Revisar fecha",
                    "nombre": "Paciente Error",
                    "rut": "333333333",
                    "fecha_derivacion": "",
                    "fecha_original": "fecha mala",
                    "edad": 45,
                    "diagnostico": "Rodilla",
                    "prioridad": "MODERADA",
                    "percapita_desde": "CESFAM",
                    "profesional": "KINESIOLOGO",
                    "observaciones": "",
                    "paciente_id": None,
                    "paciente_rut": None,
                    "paciente_nombre": None,
                    "paciente_estado": None,
                    "paciente_id_ccr": None,
                    "kine_asignado_nombre": None,
                    "requiere_revision": True,
                    "estado_revision": "PENDIENTE",
                    "resolucion": "",
                    "resuelto_en": None,
                    "resuelto_por_id": None,
                    "resuelto_por_nombre": None,
                }
            ],
        )

        listado = self.client.get("/api/importar/revision/?estado=PENDIENTE")
        self.assertEqual(listado.status_code, status.HTTP_200_OK)
        self.assertEqual(listado.data["pendientes"], 1)

        patch = self.client.patch(
            f"/api/importar/revision/{importacion.id}/0/",
            {"accion": "DESCARTAR", "resolucion": "No corresponde al corte."},
            format="json",
        )

        self.assertEqual(patch.status_code, status.HTTP_200_OK)
        importacion.refresh_from_db()
        self.assertEqual(importacion.observaciones_revision[0]["estado_revision"], "DESCARTADO")

    def test_historial_mensual_devuelve_importaciones_y_resumen_corte(self):
        importacion = ImportacionMensual.objects.create(
            archivo=ContentFile(b"test", name="historial.xlsx"),
            archivo_nombre="historial.xlsx",
            mes=5,
            anio=2026,
            mes_datos=7,
            anio_datos=2025,
            usuario=self.admin,
            total_registros=1,
            registros_importados=1,
        )
        Paciente.objects.create(
            id_ccr="HIST-1",
            fecha_derivacion=date(2025, 7, 5),
            percapita_desde="CESFAM",
            nombre="PACIENTE HISTORIAL",
            rut="555555555",
            edad=50,
            diagnostico="Lumbago",
            profesional="KINESIOLOGO",
            prioridad=Paciente.Prioridad.MODERADA,
            categoria=Paciente.Categoria.LUMBAGOS,
            importacion_origen=importacion,
        )

        response = self.client.get("/api/importar/historial/7/2025/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["periodo_label"], "Julio 2025")
        self.assertEqual(len(response.data["items"]), 1)
        self.assertEqual(response.data["pacientes_actuales_del_corte"]["pendientes"], 1)

    def test_exportar_historial_mensual_devuelve_xlsx(self):
        importacion = ImportacionMensual.objects.create(
            archivo=ContentFile(b"test", name="historial.xlsx"),
            archivo_nombre="historial.xlsx",
            mes=5,
            anio=2026,
            mes_datos=7,
            anio_datos=2025,
            usuario=self.admin,
            total_registros=1,
            registros_importados=1,
        )
        Paciente.objects.create(
            id_ccr="CCR-9999",
            fecha_derivacion=date(2025, 7, 5),
            percapita_desde="CESFAM",
            nombre="PACIENTE CORTE",
            rut="666666666",
            edad=50,
            diagnostico="Lumbago",
            profesional="KINESIOLOGO",
            prioridad=Paciente.Prioridad.MODERADA,
            categoria=Paciente.Categoria.LUMBAGOS,
            importacion_origen=importacion,
        )

        response = self.client.get("/api/importar/historial/7/2025/exportar/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], EXCEL_CONTENT_TYPE)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        self.assertEqual(ws["A1"].value, "ListaEsperaCCR")
        self.assertIn("Importación origen", [cell.value for cell in ws[7]])

    def test_vaciar_datos_ccr_no_ejecuta_sin_confirmar(self):
        usuario = Usuario.objects.create_user(
            rut="77777777-7",
            password="testpass",
            nombre="Kine Conservado",
            rol=Usuario.Rol.KINE,
        )
        Paciente.objects.create(
            id_ccr="VAC-1",
            fecha_derivacion=date(2025, 7, 5),
            nombre="PACIENTE VACIO",
            rut="777777771",
            edad=50,
            diagnostico="Lumbago",
            profesional="KINESIOLOGO",
            prioridad=Paciente.Prioridad.MODERADA,
            categoria=Paciente.Categoria.LUMBAGOS,
        )

        salida = StringIO()
        call_command("vaciar_datos_ccr", stdout=salida)

        self.assertIn("falta --confirmar", salida.getvalue())
        self.assertEqual(Paciente.objects.count(), 1)
        self.assertTrue(Usuario.objects.filter(pk=usuario.pk).exists())

    def test_vaciar_datos_ccr_borra_operativos_y_conserva_usuarios(self):
        usuario = Usuario.objects.create_user(
            rut="88888888-8",
            password="testpass",
            nombre="Kine Conservado",
            rol=Usuario.Rol.KINE,
        )
        paciente = Paciente.objects.create(
            id_ccr="VAC-2",
            fecha_derivacion=date(2025, 7, 5),
            nombre="PACIENTE VACIO",
            rut="888888881",
            edad=50,
            diagnostico="Lumbago",
            profesional="KINESIOLOGO",
            prioridad=Paciente.Prioridad.MODERADA,
            categoria=Paciente.Categoria.LUMBAGOS,
        )
        MovimientoPaciente.objects.create(
            paciente=paciente,
            usuario=usuario,
            estado_anterior=None,
            estado_nuevo=Paciente.Estado.PENDIENTE,
            notas="Movimiento de prueba",
        )
        ImportacionMensual.objects.create(
            archivo=ContentFile(b"test", name="vaciar.xlsx"),
            archivo_nombre="vaciar.xlsx",
            mes=5,
            anio=2026,
            mes_datos=7,
            anio_datos=2025,
            usuario=self.admin,
        )

        salida = StringIO()
        call_command("vaciar_datos_ccr", confirmar=True, force_production=True, stdout=salida)

        self.assertEqual(Paciente.objects.count(), 0)
        self.assertEqual(MovimientoPaciente.objects.count(), 0)
        self.assertEqual(ImportacionMensual.objects.count(), 0)
        self.assertTrue(Usuario.objects.filter(pk=usuario.pk).exists())
        self.assertIn("Datos operativos CCR eliminados", salida.getvalue())
