from datetime import date, datetime
import re
import unicodedata
import uuid

from django.utils import timezone
from openpyxl import load_workbook

from apps.pacientes.models import Paciente
from apps.pacientes.services import categoria_por_diagnostico, prioridad_normalizada
from apps.usuarios.models import Usuario

CENTROS_VALIDOS = {
    "CAR", "CST", "CCEQ", "CCE", "CES", "HT", "HH", "TMT", "FST",
    "HLH", "TMT HT", "FST HT", "CEQ", "CESFAM", "CECOSF", "SANTO",
}

MESES_IGNORAR = {"LISTAS", "INSTRUCCIONES", "REVISION", "REVISIÓN", "RESUMEN"}
MESES_VALIDOS = {
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
}

MESES_NUM = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Alias normalizados para la detección dinámica de columnas
# IMPORTANTE: Las listas van de más específico a más genérico. Se usa coincidencia EXACTA.
COLUMN_ALIASES: dict[str, list[str]] = {
    "fecha":        ["FECHA DERIV", "FECHA DERIV.", "FECHA DERIVACION",
                     "F DERIVACION", "FECHA"],
    "nombre":       ["NOMBRE", "NOMBRE COMPLETO", "PACIENTE", "NOMBRES Y APELLIDOS"],
    "rut":          ["RUT", "RUT PACIENTE", "RUN"],
    "edad":         ["EDAD", "EDAD ANOS", "ANOS"],
    "desde":        ["PERCÁPITA / DESDE", "PERCAPITA / DESDE", "PERCAPITA/DESDE",
                     "PERCAPITA DESDE", "DESDE", "PERCAPITA", "CENTRO ORIGEN"],
    "sector_oficial": [
        "SECTOR OFICIAL", "SECTOR", "Sector", "SECTOR ORIGEN",
        "SECTOR TERRITORIAL", "TERRITORIO",
    ],
    "sector_cesfam": [
        "SectorCesfam", "SECTOCESFAM", "SECTOR CESFAM", "SECTOR/CESFAM",
        "CLASIFICACION CESFAM", "CLASIFICACIÓN CESFAM", "COLOR CESFAM",
    ],
    "diagnostico":  ["DIAGNÓSTICO", "DIAGNOSTICO MEDICO", "DIAGNOSTICO", "DX"],
    "profesional":  ["PROFESIONAL DERIVADO", "PROFESIONAL DERIVACION", "PROFESIONAL",
                     "PROF DERIVADO", "KINESIOLOGO"],
    "prioridad":    ["PRIORIDAD", "GRADO PRIORIDAD", "GRADO DE PRIORIDAD", "GRADO"],
    "mayor_60":     [
        "≥60", "≥ 60", ">= 60", ">=60", ">60", "60", "MAYOR 60",
        "MAYOR60", "MAYOR O IGUAL 60", "MAYOR O IGUAL A 60",
    ],
    "observaciones":["OBSERVACIONES", "OBJETIVOS DEL TRATAMIENTO OBSERVACIONES",
                     "OBJETIVOS DEL TRATAMIENTO (OBSERVACIONES)",
                     "OBJETIVOS DEL TRATAMIENTO", "OBS", "COMENTARIOS", "NOTAS"],
    "kine_asignado": [
        "KINE ASIGNADO", "RESPONSABLE CCR", "RESPONSABLE", "KINE",
    ],
    "estado_sugerido": ["ESTADO SUGERIDO", "ESTADO", "SITUACION", "SITUACIÓN"],
    "recepcion_original": [
        "RECEPCION ORIGINAL", "RECEPCIÓN ORIGINAL", "RECEPCION", "RECEPCIÓN",
        "TOMADO POR", "INGRESA",
    ],
    "categoria_excel": [
        "CATEGORIA", "CATEGORÍA", "CATEGORIA CCR", "CATEGORÍA CCR",
        "TIPO CATEGORIA", "TIPO CATEGORÍA",
    ],
    "asignado_historico": [
        "ASIGNADO HISTORICO", "ASIGNADO HISTÓRICO", "PREORDEN",
        "PREORDEN HISTORICO", "PREORDEN HISTÓRICO",
    ],
}

KINES_ACTIVOS_CCR: dict[str, list[str]] = {
    "Mane Sáez": ["MANE", "MANE SAEZ", "MANE SÁEZ", "MARÍA ELENA SAEZ", "MARIA ELENA SAEZ", "M. SAEZ"],
    "Sebastián Campos": [
        "SEBA C", "SEBA CA", "SEBA CAMPOS", "SEBASTIAN CAMPOS", "SEBASTIÁN CAMPOS",
        "INGRESA SEBA CAMPOS", "S. CAMPOS",
    ],
    "Sebastián Salgado": [
        "SEBA SALGADO", "SEBASTIAN SALGADO", "SEBASTIÁN SALGADO", "SEBA SA", "S. SALGADO",
    ],
    "Pilar Alarcón": ["PILAR", "PILAR ALARCON", "PILAR ALARCÓN", "P. ALARCON", "P. ALARCÓN"],
    "Karen Torres": ["KAREN", "KAREN TORRES", "K. TORRES"],
}

KINES_NO_ACTIVOS = {
    "BENJA": "Benja",
    "BENJAMIN": "Benjamín",
    "BENJAMÍN": "Benjamín",
    "MARIA IGNACIA": "María Ignacia",
    "MARÍA IGNACIA": "María Ignacia",
    "MA IGNACIA": "María Ignacia",
    "M° IGNACIA": "María Ignacia",
    "M IGNACIA": "María Ignacia",
}

ESTADOS_SUGERIDOS_VALIDOS = {valor for valor, _label in Paciente.Estado.choices}

MOTIVOS_REVISION_BLOQUEANTE: tuple[tuple[str, str], ...] = (
    ("FALLECID", "Paciente marcado como fallecido."),
    ("RECHAZA", "Paciente marcado como rechaza atención."),
    ("NO ACEPTA", "Paciente marcado como rechaza atención."),
    ("YA RESUELTO", "Paciente posiblemente ya resuelto."),
    ("RESUELTO", "Paciente posiblemente ya resuelto."),
    ("ATENDIDO FUERA", "Paciente atendido fuera del flujo normal."),
    ("ATENCION EXTERNA", "Paciente atendido fuera del flujo normal."),
    ("ATENCIÓN EXTERNA", "Paciente atendido fuera del flujo normal."),
    ("PARTICULAR", "Paciente atendido fuera del flujo normal."),
)


def _cs(valor) -> str:
    if valor is None:
        return ""
    return re.sub(r"[\s\t]+", " ", str(valor)).strip()


def _normalizar(valor: str) -> str:
    texto = _cs(valor).upper()
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    # Remover puntuación al inicio/final para que "FECHA DERIV." == "FECHA DERIV"
    sin_puntuacion = re.sub(r"^[.\-_/\s]+|[.\-_/\s]+$", "", sin_tildes)
    return " ".join(sin_puntuacion.split())


normalizar_texto = _normalizar


def _normalizar_rut(valor: str) -> str:
    return re.sub(r"[^0-9Kk]", "", valor or "").upper().strip()


def _normalizar_nombre_paciente(valor: str) -> str:
    return " ".join(_cs(valor).split()).upper()


def _normalizar_percapita(valor: str) -> str:
    limpio = _cs(valor)
    normalizado = _normalizar(limpio)
    if not normalizado:
        return ""
    centros = {
        "CAR": "CESFAM DR. ALBERTO REYES",
        "CESFAM": "CESFAM DR. ALBERTO REYES",
        "ALBERTO REYES": "CESFAM DR. ALBERTO REYES",
        "CESFAM DR ALBERTO REYES": "CESFAM DR. ALBERTO REYES",
        "CESFAM DR. ALBERTO REYES": "CESFAM DR. ALBERTO REYES",
        "DR ALBERTO REYES": "CESFAM DR. ALBERTO REYES",
        "DR. ALBERTO REYES": "CESFAM DR. ALBERTO REYES",
        "HT": "HOSPITAL DE TOMÉ",
        "HOSPITAL DE TOME": "HOSPITAL DE TOMÉ",
        "HOSPITAL DE TOMÉ": "HOSPITAL DE TOMÉ",
        "HH": "HOSPITAL HIGUERAS",
        "HOSPITAL HIGUERAS": "HOSPITAL HIGUERAS",
        "FST": "FST",
        "FST HT": "FST HT",
        "TMT": "TMT",
        "TMT HT": "TMT HT",
        "CERRO ESTANQUE": "CECOSF CERRO ESTANQUE",
        "CECOSF CERRO ESTANQUE": "CECOSF CERRO ESTANQUE",
        "CCE": "CECOSF CERRO ESTANQUE",
        "CCEC": "CECOSF CERRO ESTANQUE",
        "EL SANTO": "CECOSF EL SANTO",
        "CECOSF EL SANTO": "CECOSF EL SANTO",
        "CES": "CECOSF EL SANTO",
    }
    return centros.get(normalizado, limpio)


def _normalizar_sector_oficial(valor: str) -> str:
    return _normalizar(valor)


def _normalizar_sector_cesfam(valor: str) -> tuple[str, str]:
    normalizado = _normalizar(valor)
    if not normalizado or normalizado in {"SIN SECTOR", "NO ENCONTRADO", "NO IDENTIFICADO", "NO IDENTIFICABLE"}:
        return "NO IDENTIFICABLE", "SectorCesfam vacío o no identificable."
    mapping = {
        "AZUL": "AZUL",
        "SECTOR AZUL": "AZUL",
        "ROJO": "ROJO",
        "SECTOR ROJO": "ROJO",
        "VERDE": "VERDE",
        "SECTOR VERDE": "VERDE",
        "HT": "HOSPITAL DE TOMÉ",
        "HOSPITAL DE TOME": "HOSPITAL DE TOMÉ",
        "HOSPITAL DE TOMÉ": "HOSPITAL DE TOMÉ",
        "HH": "HOSPITAL HIGUERAS",
        "HOSPITAL HIGUERAS": "HOSPITAL HIGUERAS",
        "FST": "FST",
        "FST HT": "FST HT",
        "TMT": "TMT",
        "TMT HT": "TMT HT",
        "CERRO ESTANQUE": "CECOSF CERRO ESTANQUE",
        "CECOSF CERRO ESTANQUE": "CECOSF CERRO ESTANQUE",
        "CCE": "CECOSF CERRO ESTANQUE",
        "CCEC": "CECOSF CERRO ESTANQUE",
        "EL SANTO": "CECOSF EL SANTO",
        "CECOSF EL SANTO": "CECOSF EL SANTO",
        "CES": "CECOSF EL SANTO",
    }
    return mapping.get(normalizado, normalizado), ""


def _normalizar_categoria_excel(valor: str, diagnostico: str, edad: int) -> tuple[str, str]:
    normalizada = _normalizar(valor).replace(" ", "_")
    calculada = categoria_por_diagnostico(diagnostico, edad)
    if not normalizada:
        return calculada, ""

    mapping = {
        "BORRADOR": Paciente.Categoria.BORRADOR,
        "NO_CATEGORIZADO": Paciente.Categoria.BORRADOR,
        "SIN_CATEGORIA": Paciente.Categoria.BORRADOR,
        "MAS65": Paciente.Categoria.MAS65,
        "MAYOR_65": Paciente.Categoria.MAS65,
        "MAYOR_O_IGUAL_65": Paciente.Categoria.MAS65,
        ">=65": Paciente.Categoria.MAS65,
        "OA_MENOS65": Paciente.Categoria.OA_MENOS65,
        "OA_MENOS_65": Paciente.Categoria.OA_MENOS65,
        "OSTEOARTROSIS_MENOR_65": Paciente.Categoria.OA_MENOS65,
        "HOMBROS": Paciente.Categoria.HOMBROS,
        "HOMBRO": Paciente.Categoria.HOMBROS,
        "LUMBAGOS": Paciente.Categoria.LUMBAGOS,
        "LUMBAGO": Paciente.Categoria.LUMBAGOS,
        "COLUMNA": Paciente.Categoria.LUMBAGOS,
        "SDNT": Paciente.Categoria.SDNT,
        "SDT": Paciente.Categoria.SDT,
        "OTROS_NEUROS": Paciente.Categoria.OTROS_NEUROS,
        "NEUROLOGICO": Paciente.Categoria.OTROS_NEUROS,
        "NEUROLÓGICO": Paciente.Categoria.OTROS_NEUROS,
        "AATT": Paciente.Categoria.AATT,
        "ACCIDENTE_TRABAJO": Paciente.Categoria.AATT,
        "DUPLA": Paciente.Categoria.DUPLA,
    }
    categoria = mapping.get(normalizada)
    if categoria:
        return categoria, ""
    return calculada, f"Categoría inválida: {valor}. Se calculó automáticamente."


def _leer_celda(ws, row_idx: int, columnas: dict, campo: str) -> str:
    if campo not in columnas:
        return ""
    return _cs(ws.cell(row_idx, columnas[campo]).value)


def _contiene_alias(texto_norm: str, alias: str) -> bool:
    alias_norm = _normalizar(alias)
    if not alias_norm:
        return False
    patron = rf"(?<![A-Z0-9]){re.escape(alias_norm)}(?![A-Z0-9])"
    return bool(re.search(patron, texto_norm))


def _usuarios_activos_kine_por_canonico() -> dict[str, Usuario]:
    usuarios = Usuario.objects.filter(rol=Usuario.Rol.KINE, is_active=True)
    indice: dict[str, Usuario] = {}
    for usuario in usuarios:
        nombre_norm = _normalizar(usuario.nombre)
        for canonico, aliases in KINES_ACTIVOS_CCR.items():
            posibles = [canonico, *aliases]
            if any(_normalizar(alias) == nombre_norm for alias in posibles):
                indice.setdefault(canonico, usuario)
    return indice


def _detectar_kines_en_texto(texto: str) -> tuple[set[str], set[str]]:
    texto_norm = _normalizar(texto)
    activos: set[str] = set()
    no_activos: set[str] = set()
    if not texto_norm:
        return activos, no_activos

    for canonico, aliases in KINES_ACTIVOS_CCR.items():
        if any(_contiene_alias(texto_norm, alias) for alias in aliases):
            activos.add(canonico)

    for alias, label in KINES_NO_ACTIVOS.items():
        if _contiene_alias(texto_norm, alias):
            no_activos.add(label)

    return activos, no_activos


def _resolver_kine(
    *,
    valores_explicitos: list[tuple[str, str]],
    usuarios_kine: dict[str, Usuario],
) -> dict:
    fuentes = [(fuente, valor) for fuente, valor in valores_explicitos if valor]
    fuentes_busqueda = fuentes

    detectados_activos: set[str] = set()
    detectados_no_activos: set[str] = set()
    fuente_detectada = ""
    for fuente, valor in fuentes_busqueda:
        activos, no_activos = _detectar_kines_en_texto(valor)
        if activos or no_activos:
            fuente_detectada = fuente_detectada or fuente
        detectados_activos.update(activos)
        detectados_no_activos.update(no_activos)

    total_detectados = len(detectados_activos) + len(detectados_no_activos)
    if total_detectados > 1:
        nombres = sorted([*detectados_activos, *detectados_no_activos])
        return {
            "usuario": None,
            "kine_detectado": ", ".join(nombres),
            "fuente": fuente_detectada,
            "motivo_revision": "Kine ambiguo detectado: " + ", ".join(nombres),
        }

    if detectados_no_activos:
        nombre = next(iter(detectados_no_activos))
        return {
            "usuario": None,
            "kine_detectado": nombre,
            "fuente": fuente_detectada,
            "motivo_revision": f"Kine no activo detectado: {nombre}.",
        }

    if detectados_activos:
        canonico = next(iter(detectados_activos))
        usuario = usuarios_kine.get(canonico)
        if usuario:
            return {
                "usuario": usuario,
                "kine_detectado": usuario.nombre,
                "fuente": fuente_detectada,
                "motivo_revision": "",
            }
        return {
            "usuario": None,
            "kine_detectado": canonico,
            "fuente": fuente_detectada,
            "motivo_revision": f"Kine CCR detectado pero no existe como usuario activo: {canonico}.",
        }

    for fuente, valor in fuentes:
        valor_norm = _normalizar(valor)
        if valor_norm and valor_norm not in {"SIN", "SIN ASIGNAR", "NO", "N/A", "NA", "PENDIENTE", "S/R"}:
            return {
                "usuario": None,
                "kine_detectado": valor,
                "fuente": fuente,
                "motivo_revision": f"Kine no confirmado detectado: {valor}.",
            }

    return {"usuario": None, "kine_detectado": "", "fuente": "", "motivo_revision": ""}


def _normalizar_estado_sugerido(valor: str) -> tuple[str, str]:
    estado = _normalizar(valor).replace(" ", "_")
    if not estado:
        return Paciente.Estado.PENDIENTE, ""
    if estado in ESTADOS_SUGERIDOS_VALIDOS:
        return estado, ""
    return Paciente.Estado.PENDIENTE, f"Estado sugerido inválido: {valor}."


def _normalizar_asignado_historico(valor: str) -> bool:
    return _normalizar(valor) in {"SI", "S", "1", "TRUE", "X"}


def _motivos_revision_bloqueante(*valores: str) -> list[str]:
    texto = _normalizar(" ".join(valor for valor in valores if valor))
    motivos: list[str] = []
    for marcador, motivo in MOTIVOS_REVISION_BLOQUEANTE:
        if _normalizar(marcador) in texto and motivo not in motivos:
            motivos.append(motivo)
    return motivos


def _armar_observaciones(
    *,
    base: str,
    sector_oficial: str,
    sector_cesfam: str,
    recepcion_original: str,
    kine_detectado: str,
    estado_sugerido: str,
    categoria: str,
    motivos_revision: list[str],
    asignado_historico: bool = False,
) -> str:
    partes: list[str] = []
    if base:
        partes.append(f"Observación original: {base}")
    if sector_oficial:
        partes.append(f"Sector oficial: {sector_oficial}")
    if sector_cesfam:
        partes.append(f"SectorCesfam: {sector_cesfam}")
    if recepcion_original:
        partes.append(f"Recepción original: {recepcion_original}")
    if kine_detectado:
        partes.append(f"Responsable detectado: {kine_detectado}")
    if estado_sugerido:
        partes.append(f"Estado sugerido: {estado_sugerido}")
    if categoria:
        partes.append(f"Categoría: {categoria}")
    if asignado_historico:
        partes.append("Registro incluido en preorden histórico CCR")
    if motivos_revision:
        partes.append("Revisión: " + " ".join(motivos_revision))
    return ". ".join(partes) if partes else ""


def _parsear_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if valor is None:
        return None

    texto = str(valor).strip()
    # Buscar primero formato con año de 4 dígitos
    match4 = re.search(r"(\d{1,2}[./-]\d{1,2}[./-]\d{4}|\d{4}-\d{1,2}-\d{1,2})", texto)
    if match4:
        texto = match4.group(1)
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue

    # Buscar formato con año de 2 dígitos (ej: 30-09-24)
    match2 = re.search(r"(\d{1,2}[./-]\d{1,2}[./-]\d{2})$", texto)
    if match2:
        texto2 = match2.group(1)
        for fmt in ("%d.%m.%y", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(texto2, fmt).date()
            except ValueError:
                continue

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue

    try:
        numero = float(texto)
        from openpyxl.utils.datetime import from_excel
        convertido = from_excel(numero)
        if isinstance(convertido, datetime):
            return convertido.date()
        if isinstance(convertido, date):
            return convertido
    except Exception:
        pass

    return None


def _mes_desde_hoja(sheet_name: str) -> str | None:
    nombre = _normalizar(sheet_name)
    for mes in MESES_VALIDOS:
        if re.search(rf"\b{re.escape(mes)}\b", nombre):
            return mes
    return None


def _detectar_columnas_dinamicas(ws) -> dict[str, int] | None:
    """
    Busca en las primeras filas de la hoja encabezados reconocibles, incluso cuando
    vienen repartidos en dos filas como las planillas "Derivaciones CCR",
    y devuelve un mapeo campo->número_de_columna (1-indexed).
    Usa coincidencia EXACTA con los aliases normalizados para evitar falsos positivos.
    Devuelve None si no puede identificar las columnas mínimas.
    """
    # Pre-normalizar todos los aliases
    aliases_norm: dict[str, list[str]] = {
        field: [_normalizar(a) for a in aliases]
        for field, aliases in COLUMN_ALIASES.items()
    }

    max_col = min(max(ws.max_column, 22), 80)
    filas_normalizadas: dict[int, dict[int, str]] = {}

    for row_idx in range(1, min(ws.max_row, 12) + 1):
        row_vals: dict[int, str] = {}
        all_empty = True
        for col in range(1, max_col + 1):
            raw = ws.cell(row_idx, col).value
            if raw is not None:
                all_empty = False
            norm = _normalizar(_cs(raw))
            row_vals[col] = norm

        if not all_empty:
            filas_normalizadas[row_idx] = row_vals

    for row_idx in sorted(filas_normalizadas):
        for span in (1, 2, 3):
            rows = [
                filas_normalizadas.get(idx, {})
                for idx in range(row_idx, row_idx + span)
            ]
            if not any(rows):
                continue

            mapping: dict[str, int] = {}
            for field, norm_aliases in aliases_norm.items():
                for row_vals in rows:
                    for col, cell_norm in row_vals.items():
                        if not cell_norm:
                            continue
                        # Coincidencia EXACTA: la celda normalizada debe estar en los aliases
                        if cell_norm in norm_aliases:
                            if field not in mapping:
                                mapping[field] = col
                            break
                    if field in mapping:
                        break

            # Necesitamos al menos nombre, rut y diagnóstico para considerar válida la hoja
            if all(k in mapping for k in ("nombre", "rut", "diagnostico")):
                return {"header_row": row_idx + span - 1, **mapping}

    return None


def _dup_key(rut: str, fecha_derivacion: date, diagnostico: str) -> tuple[str, date, str]:
    return (rut, fecha_derivacion, (diagnostico or "").upper().strip())


def _fecha_preview(fecha_derivacion: date | None) -> str:
    return fecha_derivacion.strftime("%d/%m/%Y") if fecha_derivacion else ""


def _periodo_preview(fecha_derivacion: date | None) -> str | None:
    return fecha_derivacion.strftime("%Y-%m") if fecha_derivacion else None


def _valor_preview(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def _registro_preview(
    *,
    hoja: str,
    fila: int,
    nombre: str,
    rut: str,
    fecha_derivacion: date | None,
    edad: int,
    diagnostico: str,
    prioridad: str,
    percapita_desde: str,
    sector_oficial: str,
    sector_cesfam: str,
    profesional: str,
    observaciones: str,
    mayor_60: bool,
    categoria: str,
    kine_asignado: str = "",
    estado_sugerido: str = "",
    recepcion_original: str = "",
    asignado_historico: bool = False,
    motivo_revision: str = "",
    advertencias: list[str] | None = None,
    es_duplicado: bool = False,
    error: str = "",
) -> dict:
    advertencias = advertencias or []
    if error:
        estado = "ERROR"
        tipo_revision = "ERROR"
    elif es_duplicado:
        estado = "DUPLICADO"
        tipo_revision = "RECURRENTE"
    elif motivo_revision or advertencias:
        estado = "REVISION"
        tipo_revision = "ADVERTENCIA"
    else:
        estado = "OK"
        tipo_revision = ""

    return {
        "hoja": hoja,
        "fila": fila,
        "nombre": nombre,
        "rut": rut,
        "fecha_derivacion": _fecha_preview(fecha_derivacion),
        "edad": edad,
        "diagnostico": diagnostico,
        "prioridad": prioridad,
        "percapita_desde": percapita_desde,
        "sector_oficial": sector_oficial,
        "sector_cesfam": sector_cesfam,
        "profesional": profesional,
        "observaciones": observaciones,
        "mayor_60": mayor_60,
        "categoria": categoria,
        "kine_asignado": kine_asignado,
        "estado_sugerido": estado_sugerido,
        "recepcion_original": recepcion_original,
        "asignado_historico": asignado_historico,
        "motivo_revision": motivo_revision,
        "advertencias": advertencias,
        "es_duplicado": es_duplicado,
        "estado": estado,
        "error": error,
        "tipo_revision": tipo_revision,
    }


def _procesar_hoja(
    ws,
    sheet_name: str,
    columnas: dict,
    header_row: int,
    existentes: set,
    ruts_existentes: set,
    ruts_vistos_importacion: set,
    duplicados_vistos: set,
    usuarios_kine: dict[str, Usuario],
    mes_objetivo: int | None = None,
    anio_objetivo: int | None = None,
) -> dict:
    """
    Procesa filas de una hoja usando el mapeo dinámico de columnas.
    Retorna dict con registros, pacientes_a_crear, errores, total, duplicados, fechas.
    """
    registros = []
    pacientes_a_crear = []
    errores = []
    total = 0
    duplicados = 0
    fechas_hoja: list[date] = []

    campos_lectura = [
        "fecha", "nombre", "rut", "edad", "desde", "sector_oficial", "sector_cesfam",
        "diagnostico", "profesional",
        "prioridad", "mayor_60", "observaciones", "kine_asignado",
        "estado_sugerido", "recepcion_original", "categoria_excel", "asignado_historico",
    ]

    for row_idx in range(header_row + 1, ws.max_row + 1):
        valores = {campo: _leer_celda(ws, row_idx, columnas, campo) for campo in campos_lectura}
        if not any(valores.values()):
            continue

        nombre = _normalizar_nombre_paciente(valores["nombre"])
        rut = _normalizar_rut(valores["rut"])
        diagnostico = _cs(valores["diagnostico"])
        profesional = _cs(valores["profesional"])
        prioridad = prioridad_normalizada(valores["prioridad"] or "")
        observaciones_base = _cs(valores["observaciones"])
        desde = _normalizar_percapita(valores["desde"])
        sector_oficial = _normalizar_sector_oficial(valores["sector_oficial"])
        sector_cesfam, sector_cesfam_motivo = _normalizar_sector_cesfam(valores["sector_cesfam"])
        if not valores["sector_cesfam"] and desde:
            sector_cesfam = desde
            sector_cesfam_motivo = ""
        recepcion_original = _cs(valores["recepcion_original"])
        categoria_excel = _cs(valores["categoria_excel"])
        asignado_historico = _normalizar_asignado_historico(valores["asignado_historico"])
        estado_sugerido_raw = _cs(valores["estado_sugerido"])
        estado_sugerido, estado_motivo = _normalizar_estado_sugerido(estado_sugerido_raw)
        fecha_derivacion = None
        edad = 0
        mayor_60 = False
        categoria = categoria_por_diagnostico(diagnostico, 0)
        motivos_revision: list[str] = []

        nombre_norm = _normalizar(nombre)
        if nombre_norm in {"NOMBRE", "N°", "NOMBRE COMPLETO", "PACIENTE"}:
            continue

        total += 1

        try:
            if not nombre:
                raise ValueError("Nombre vacío")
            if not rut:
                raise ValueError("RUT vacío")

            fecha_raw = ws.cell(row_idx, columnas["fecha"]).value if "fecha" in columnas else None
            if "fecha" in columnas:
                fecha_derivacion = _parsear_fecha(fecha_raw)
            if fecha_derivacion is None:
                raise ValueError(
                    f"Fecha inválida o ausente: {ws.cell(row_idx, columnas.get('fecha', 1)).value!r}"
                )
            if (
                mes_objetivo
                and anio_objetivo
                and (fecha_derivacion.month != mes_objetivo or fecha_derivacion.year != anio_objetivo)
            ):
                total -= 1
                continue
            fechas_hoja.append(fecha_derivacion)

            if "edad" in columnas:
                edad_valor = ws.cell(row_idx, columnas["edad"]).value
                try:
                    edad = int(float(str(edad_valor))) if edad_valor not in (None, "") else 0
                except (ValueError, TypeError):
                    edad = 0
            mayor_60 = edad >= 60
            if "mayor_60" in columnas:
                val_m60 = _normalizar(_cs(ws.cell(row_idx, columnas["mayor_60"]).value))
                if val_m60 in {"SI", "S", "1", "TRUE", "X"}:
                    mayor_60 = True
                elif val_m60 in {"NO", "N", "0", "FALSE", ""}:
                    mayor_60 = edad >= 60
            categoria, categoria_motivo = _normalizar_categoria_excel(categoria_excel, diagnostico, edad)

            if not diagnostico:
                raise ValueError("Diagnóstico vacío")

            if estado_motivo:
                motivos_revision.append(estado_motivo)
            if not sector_oficial:
                motivos_revision.append("Sector oficial vacío.")
            if sector_cesfam_motivo:
                motivos_revision.append(sector_cesfam_motivo)
            if categoria_motivo:
                motivos_revision.append(categoria_motivo)

            kine_info = _resolver_kine(
                valores_explicitos=[
                    ("KINE ASIGNADO", valores["kine_asignado"]),
                ],
                usuarios_kine=usuarios_kine,
            )
            if kine_info["motivo_revision"]:
                motivos_revision.append(kine_info["motivo_revision"])

            motivos_revision.extend(_motivos_revision_bloqueante(observaciones_base, diagnostico))

            observaciones = _armar_observaciones(
                base=observaciones_base,
                sector_oficial=sector_oficial,
                sector_cesfam=sector_cesfam,
                recepcion_original=recepcion_original,
                kine_detectado=kine_info["kine_detectado"],
                estado_sugerido=estado_sugerido_raw,
                categoria=categoria,
                motivos_revision=motivos_revision,
                asignado_historico=asignado_historico,
            )

            dup_key = _dup_key(rut, fecha_derivacion, diagnostico)
            rut_recurrente = rut in ruts_existentes or rut in ruts_vistos_importacion
            if dup_key in duplicados_vistos or dup_key in existentes or rut_recurrente:
                duplicados += 1
                registros.append(_registro_preview(
                    hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                    fecha_derivacion=fecha_derivacion, edad=edad,
                    diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                    sector_oficial=sector_oficial, sector_cesfam=sector_cesfam,
                    profesional=profesional, observaciones=observaciones,
                    mayor_60=mayor_60, categoria=categoria,
                    kine_asignado=kine_info["kine_detectado"],
                    estado_sugerido=estado_sugerido,
                    recepcion_original=recepcion_original,
                    asignado_historico=asignado_historico,
                    motivo_revision=" ".join(motivos_revision),
                    advertencias=motivos_revision,
                    es_duplicado=True,
                ))
                continue

            duplicados_vistos.add(dup_key)
            ruts_vistos_importacion.add(rut)
            fecha_cambio_estado = timezone.now() if estado_sugerido != Paciente.Estado.PENDIENTE else None
            paciente = Paciente(
                id_ccr=f"TMP{uuid.uuid4().hex[:9].upper()}",
                fecha_derivacion=fecha_derivacion,
                percapita_desde=desde,
                sector_oficial=sector_oficial,
                sector_cesfam=sector_cesfam,
                nombre=nombre,
                rut=rut,
                edad=edad,
                diagnostico=diagnostico,
                profesional=profesional,
                prioridad=prioridad,
                categoria=categoria,
                mayor_60=mayor_60,
                observaciones=observaciones,
                kine_asignado=kine_info["usuario"],
                estado=estado_sugerido,
                fecha_cambio_estado=fecha_cambio_estado,
                asignado_historico=asignado_historico,
            )
            paciente._importacion_estado_sugerido = estado_sugerido
            paciente._importacion_kine_detectado = kine_info["kine_detectado"]
            paciente._importacion_motivo_revision = " ".join(motivos_revision)
            paciente._importacion_asignado_historico = asignado_historico
            pacientes_a_crear.append(paciente)
            registros.append(_registro_preview(
                hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                fecha_derivacion=fecha_derivacion, edad=edad,
                diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                sector_oficial=sector_oficial, sector_cesfam=sector_cesfam,
                profesional=profesional, observaciones=observaciones,
                mayor_60=mayor_60, categoria=categoria,
                kine_asignado=kine_info["kine_detectado"],
                estado_sugerido=estado_sugerido,
                recepcion_original=recepcion_original,
                asignado_historico=asignado_historico,
                motivo_revision=" ".join(motivos_revision),
                advertencias=motivos_revision,
            ))
        except Exception as exc:
            error_msg = str(exc)
            observaciones = _armar_observaciones(
                base=observaciones_base,
                sector_oficial=sector_oficial,
                sector_cesfam=sector_cesfam,
                recepcion_original=recepcion_original,
                kine_detectado="",
                estado_sugerido=estado_sugerido_raw,
                categoria=categoria,
                motivos_revision=motivos_revision,
                asignado_historico=asignado_historico,
            )
            errores.append({
                "hoja": sheet_name,
                "fila": row_idx,
                "motivo": error_msg,
                "nombre": nombre,
                "rut": rut,
                "fecha_derivacion": _fecha_preview(fecha_derivacion),
                "fecha_original": _valor_preview(ws.cell(row_idx, columnas["fecha"]).value) if "fecha" in columnas else "",
                "edad": edad,
                "diagnostico": diagnostico,
                "prioridad": prioridad,
                "percapita_desde": desde,
                "sector_oficial": sector_oficial,
                "sector_cesfam": sector_cesfam,
                "profesional": profesional,
                "observaciones": observaciones,
                "categoria": categoria,
                "kine_asignado": "",
                "estado_sugerido": estado_sugerido,
                "recepcion_original": recepcion_original,
                "asignado_historico": asignado_historico,
            })
            registros.append(_registro_preview(
                hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                fecha_derivacion=fecha_derivacion, edad=edad,
                diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                sector_oficial=sector_oficial, sector_cesfam=sector_cesfam,
                profesional=profesional, observaciones=observaciones,
                mayor_60=mayor_60, categoria=categoria,
                estado_sugerido=estado_sugerido,
                recepcion_original=recepcion_original,
                asignado_historico=asignado_historico,
                motivo_revision=" ".join(motivos_revision),
                advertencias=motivos_revision,
                error=error_msg,
            ))

    return {
        "registros": registros,
        "pacientes_a_crear": pacientes_a_crear,
        "errores": errores,
        "total": total,
        "duplicados": duplicados,
        "fechas_hoja": fechas_hoja,
    }


def _mes_desde_fechas(fechas: list[date]) -> str | None:
    """Detecta el mes dominante de una lista de fechas de derivación."""
    if not fechas:
        return None
    conteo: dict[int, int] = {}
    for f in fechas:
        conteo[f.month] = conteo.get(f.month, 0) + 1
    mes_num = max(conteo, key=lambda k: conteo[k])
    # Invertir MESES_NUM
    for nombre, num in MESES_NUM.items():
        if num == mes_num:
            return nombre
    return None


def _procesar_derivaciones(
    archivo,
    mes_objetivo: int | None = None,
    anio_objetivo: int | None = None,
) -> dict:
    wb = load_workbook(filename=archivo, data_only=True)

    total = 0
    duplicados = 0
    errores: list[dict] = []
    registros: list[dict] = []
    pacientes_a_crear: list[Paciente] = []
    meses_detectados: dict[str, int] = {}
    periodos_detectados: dict[str, int] = {}
    duplicados_vistos: set[tuple[str, date, str]] = set()
    ruts_vistos_importacion: set[str] = set()
    meses_ignorar = {_normalizar(m) for m in MESES_IGNORAR}
    existentes = {
        _dup_key(rut, fecha_derivacion, diagnostico)
        for rut, fecha_derivacion, diagnostico in Paciente.objects.values_list(
            "rut", "fecha_derivacion", "diagnostico"
        )
    }
    ruts_existentes = set(Paciente.objects.values_list("rut", flat=True))
    usuarios_kine = _usuarios_activos_kine_por_canonico()

    for sheet_name in wb.sheetnames:
        normalizado = _normalizar(sheet_name)
        if normalizado in meses_ignorar:
            continue

        ws = wb[sheet_name]

        # Intentar detectar columnas dinámicamente
        col_info = _detectar_columnas_dinamicas(ws)
        if col_info is None:
            # Hoja sin encabezados reconocibles → saltar silenciosamente
            continue

        header_row = col_info.pop("header_row")
        columnas = col_info

        resultado_hoja = _procesar_hoja(
            ws, sheet_name, columnas, header_row,
            existentes,
            ruts_existentes,
            ruts_vistos_importacion,
            duplicados_vistos,
            usuarios_kine,
            mes_objetivo=mes_objetivo,
            anio_objetivo=anio_objetivo,
        )

        total += resultado_hoja["total"]
        duplicados += resultado_hoja["duplicados"]
        errores.extend(resultado_hoja["errores"])
        registros.extend(resultado_hoja["registros"])
        pacientes_a_crear.extend(resultado_hoja["pacientes_a_crear"])
        registros_hoja = resultado_hoja["total"]

        # Determinar el mes de la hoja:
        # 1. Por nombre de hoja (ENERO, FEBRERO, etc.) → todos los registros van a ese mes
        # 2. Archivo base mezclado → distribuir por mes real de derivación
        mes_hoja = _mes_desde_hoja(sheet_name)
        if mes_hoja:
            # Hoja de un mes fijo: sumar todos los registros leídos, incluyendo recurrentes.
            meses_detectados[mes_hoja] = meses_detectados.get(mes_hoja, 0) + registros_hoja
        else:
            # Archivo base mezclado: distribuir cada paciente por su mes real
            for f in resultado_hoja["fechas_hoja"]:
                mes_nombre = None
                for nm, num in MESES_NUM.items():
                    if num == f.month:
                        mes_nombre = nm
                        break
                if mes_nombre:
                    meses_detectados[mes_nombre] = meses_detectados.get(mes_nombre, 0) + 1

        for f in resultado_hoja["fechas_hoja"]:
            periodo = _periodo_preview(f)
            if periodo:
                periodos_detectados[periodo] = periodos_detectados.get(periodo, 0) + 1

    return {
        "total": total,
        "duplicados": duplicados,
        "errores": errores,
        "registros": registros,
        "pacientes": pacientes_a_crear,
        "meses_detectados": meses_detectados,
        "periodos_detectados": periodos_detectados,
    }


def _resumen_registros_oficiales(registros: list[dict]) -> dict:
    registros_validos = [
        registro
        for registro in registros
        if registro.get("estado") not in {"ERROR", "DUPLICADO"}
    ]
    conteo_sector_cesfam: dict[str, int] = {}
    conteo_categoria: dict[str, int] = {}
    conteo_kine_asignado: dict[str, int] = {}
    ingresados = 0
    pendientes = 0
    asignado_historico_total = 0
    con_kine = 0

    for registro in registros_validos:
        estado_sugerido = registro.get("estado_sugerido") or Paciente.Estado.PENDIENTE
        if estado_sugerido == Paciente.Estado.INGRESADO:
            ingresados += 1
        if estado_sugerido == Paciente.Estado.PENDIENTE:
            pendientes += 1
        if registro.get("asignado_historico"):
            asignado_historico_total += 1

        kine = (registro.get("kine_asignado") or "").strip()
        if kine:
            con_kine += 1
            conteo_kine_asignado[kine] = conteo_kine_asignado.get(kine, 0) + 1

        sector_cesfam = registro.get("sector_cesfam") or "Sin dato"
        categoria = registro.get("categoria") or "Sin dato"
        conteo_sector_cesfam[sector_cesfam] = conteo_sector_cesfam.get(sector_cesfam, 0) + 1
        conteo_categoria[categoria] = conteo_categoria.get(categoria, 0) + 1

    return {
        "con_kine_asignado": con_kine,
        "sin_kine_asignado": max(0, len(registros_validos) - con_kine),
        "ingresados": ingresados,
        "pendientes": pendientes,
        "asignado_historico": asignado_historico_total,
        "conteo_sector_cesfam": conteo_sector_cesfam,
        "conteo_categoria": conteo_categoria,
        "conteo_kine_asignado": conteo_kine_asignado,
    }


def previsualizar_derivaciones(
    archivo,
    mes_objetivo: int | None = None,
    anio_objetivo: int | None = None,
) -> dict:
    procesado = _procesar_derivaciones(
        archivo,
        mes_objetivo=mes_objetivo,
        anio_objetivo=anio_objetivo,
    )
    errores_count = len(procesado["errores"])
    resumen_oficial = _resumen_registros_oficiales(procesado["registros"])
    return {
        "total": procesado["total"],
        "validos": len(procesado["pacientes"]),
        "nuevos": len(procesado["pacientes"]),
        "duplicados": procesado["duplicados"],
        "recurrentes": procesado["duplicados"],
        "errores": procesado["errores"],
        "errores_count": errores_count,
        "registros": procesado["registros"],
        "meses_detectados": procesado["meses_detectados"],
        "periodos_detectados": procesado["periodos_detectados"],
        **resumen_oficial,
    }


def parsear_derivaciones(
    archivo,
    mes_objetivo: int | None = None,
    anio_objetivo: int | None = None,
) -> dict:
    procesado = _procesar_derivaciones(
        archivo,
        mes_objetivo=mes_objetivo,
        anio_objetivo=anio_objetivo,
    )
    errores_count = len(procesado["errores"])
    resumen_oficial = _resumen_registros_oficiales(procesado["registros"])
    return {
        "total": procesado["total"],
        "importados": 0,  # La vista completa este valor después de crear pacientes.
        "nuevos": len(procesado["pacientes"]),
        "duplicados": procesado["duplicados"],
        "recurrentes": procesado["duplicados"],
        "errores": procesado["errores"],
        "errores_count": errores_count,
        "registros": procesado["registros"],
        "meses_detectados": procesado["meses_detectados"],
        "periodos_detectados": procesado["periodos_detectados"],
        "pacientes": procesado["pacientes"],
        **resumen_oficial,
    }


normalizar_texto = _normalizar
parsear_fecha = _parsear_fecha
