from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from apps.pacientes.models import Paciente
from apps.usuarios.models import Usuario


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ReportesOperativosTests(APITestCase):
    def setUp(self):
        self.kine = Usuario.objects.create_user(
            rut="11111111-1",
            password="testpass",
            nombre="Kine Uno",
            rol=Usuario.Rol.KINE,
        )
        self.kine_sin_pacientes = Usuario.objects.create_user(
            rut="22222222-2",
            password="testpass",
            nombre="Kine Cero",
            rol=Usuario.Rol.KINE,
        )
        self.admin = Usuario.objects.create_user(
            rut="33333333-3",
            password="testpass",
            nombre="Admin Reportes",
            rol=Usuario.Rol.ADMIN,
        )
        self.client.force_authenticate(self.admin)
        self.counter = 0

    def crear_paciente(self, **overrides):
        self.counter += 1
        data = {
            "fecha_derivacion": date(2025, 7, 1),
            "percapita_desde": "CESFAM",
            "nombre": f"Paciente Reporte {self.counter}",
            "rut": f"800000{self.counter:02d}-{self.counter % 10}",
            "edad": 50,
            "diagnostico": "Lumbago",
            "profesional": "KINESIOLOGO",
            "prioridad": Paciente.Prioridad.MODERADA,
            "categoria": Paciente.Categoria.LUMBAGOS,
            "kine_asignado": self.kine,
            "telefono": "123456789",
            "telefono_recados": "",
        }
        data.update(overrides)
        return Paciente.objects.create(**data)

    def test_resumen_mensual_separa_corte_y_actividad(self):
        self.crear_paciente(
            estado=Paciente.Estado.PENDIENTE,
            sector_cesfam="AZUL",
            sector_oficial="CENTENARIO",
        )
        self.crear_paciente(estado=Paciente.Estado.RESCATE)
        self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            fecha_ingreso=date(2025, 7, 6),
        )
        self.crear_paciente(
            estado=Paciente.Estado.ALTA_MEDICA,
            fecha_egreso=date(2025, 7, 20),
        )
        self.crear_paciente(
            estado=Paciente.Estado.EGRESO_ADMINISTRATIVO,
            fecha_egreso=date(2025, 7, 25),
        )
        self.crear_paciente(
            fecha_derivacion=date(2025, 8, 1),
            estado=Paciente.Estado.PENDIENTE,
        )

        response = self.client.get("/api/reportes/resumen/?mes=7&anio=2025")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["periodo_label"], "Julio 2025")
        self.assertEqual(response.data["corte"]["total_derivados"], 5)
        self.assertEqual(response.data["corte"]["pendientes"], 1)
        self.assertEqual(response.data["corte"]["rescate"], 1)
        self.assertEqual(response.data["corte"]["ingresados_actuales"], 1)
        self.assertEqual(response.data["corte"]["altas_medicas"], 1)
        self.assertEqual(response.data["corte"]["egresos_administrativos"], 1)
        self.assertEqual(response.data["actividad_mes"]["ingresos"], 1)
        self.assertEqual(response.data["actividad_mes"]["egresos_total"], 2)
        self.assertEqual(response.data["actividad_mes"]["altas_medicas"], 1)
        self.assertEqual(response.data["actividad_mes"]["egresos_administrativos"], 1)
        self.assertEqual(response.data["actividad_mes"]["promedio_dias_hasta_ingreso"], 5)
        por_sector_cesfam = {
            item["label"]: item["total"] for item in response.data["por_sector_cesfam"]
        }
        por_sector_oficial = {
            item["label"]: item["total"] for item in response.data["por_sector_oficial"]
        }
        por_diagnostico = {
            item["label"]: item["total"] for item in response.data["por_diagnostico"]
        }
        self.assertEqual(por_sector_cesfam["AZUL"], 1)
        self.assertEqual(por_sector_oficial["CENTENARIO"], 1)
        self.assertEqual(por_diagnostico["Lumbago"], 5)

    def test_por_responsable_incluye_kines_y_sin_responsable(self):
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE)
        self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            fecha_ingreso=date(2025, 7, 10),
        )
        self.crear_paciente(
            estado=Paciente.Estado.ABANDONO,
            fecha_egreso=date(2025, 7, 22),
        )
        self.crear_paciente(
            estado=Paciente.Estado.RESCATE,
            kine_asignado=None,
        )

        response = self.client.get("/api/reportes/por-responsable/?mes=7&anio=2025")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        responsables = {
            item["responsable_id"]: item for item in response.data["responsables"]
        }
        self.assertIn(self.kine.id, responsables)
        self.assertIn(self.kine_sin_pacientes.id, responsables)
        self.assertEqual(responsables[self.kine.id]["total_asignados_corte"], 3)
        self.assertEqual(responsables[self.kine.id]["ingresos_mes"], 1)
        self.assertEqual(responsables[self.kine.id]["egresos_mes"], 1)
        self.assertEqual(responsables[self.kine.id]["abandonos_mes"], 1)
        self.assertEqual(responsables[self.kine_sin_pacientes.id]["total_asignados_corte"], 0)
        self.assertEqual(response.data["sin_responsable"]["total_corte"], 1)
        self.assertEqual(response.data["sin_responsable"]["rescate"], 1)

    def test_serie_mensual_devuelve_12_meses_y_no_falla_sin_datos(self):
        self.crear_paciente(fecha_derivacion=date(2025, 1, 5), estado=Paciente.Estado.PENDIENTE)
        self.crear_paciente(
            fecha_derivacion=date(2025, 1, 1),
            fecha_ingreso=date(2025, 2, 1),
            estado=Paciente.Estado.INGRESADO,
        )
        self.crear_paciente(
            fecha_derivacion=date(2024, 12, 1),
            fecha_egreso=date(2025, 2, 15),
            estado=Paciente.Estado.DERIVADO,
        )

        response = self.client.get("/api/reportes/serie-mensual/?anio=2025")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["meses"]), 12)
        enero = response.data["meses"][0]
        febrero = response.data["meses"][1]
        self.assertEqual(enero["total_derivados"], 2)
        self.assertEqual(febrero["ingresos"], 1)
        self.assertEqual(febrero["egresos_total"], 1)

    def test_permisos_reportes(self):
        autenticado = self.client.get("/api/reportes/resumen/?mes=7&anio=2025")
        self.assertEqual(autenticado.status_code, status.HTTP_200_OK)

        self.client.force_authenticate(user=None)
        no_autenticado = self.client.get("/api/reportes/resumen/?mes=7&anio=2025")
        self.assertIn(
            no_autenticado.status_code,
            {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN},
        )

    def test_exportar_por_responsable_devuelve_xlsx(self):
        self.kine.nombre = "=Responsable"
        self.kine.save(update_fields=["nombre"])
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE)

        response = self.client.get("/api/reportes/por-responsable/exportar/?mes=7&anio=2025")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], EXCEL_CONTENT_TYPE)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        self.assertEqual(ws["A1"].value, "Reporte por responsable CCR")
        self.assertIn("Responsable CCR", [cell.value for cell in ws[5]])
        self.assertEqual(ws["A6"].value, "'=Responsable")
