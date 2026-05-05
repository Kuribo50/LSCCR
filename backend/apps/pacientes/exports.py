from datetime import date
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

from .models import Paciente


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORMULA_PREFIXES = ("=", "+", "-", "@")

PRIORIDAD_FILL = {
    Paciente.Prioridad.ALTA: "FEE2E2",
    Paciente.Prioridad.MEDIANA: "FEF3C7",
    Paciente.Prioridad.MODERADA: "E7F3EC",
    Paciente.Prioridad.LICENCIA_MEDICA: "E0F2FE",
}


def excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def sanitizar_celda_excel(valor):
    if isinstance(valor, str) and valor.lstrip().startswith(FORMULA_PREFIXES):
        return f"'{valor}"
    return valor


def _dias_en_lista(paciente: Paciente, hoy: date | None = None) -> int:
    hoy = hoy or timezone.localdate()
    return max(0, (hoy - paciente.fecha_derivacion).days)


def _fecha(valor):
    if not valor:
        return ""
    return valor


def _fecha_hora(valor):
    if not valor:
        return ""
    return timezone.localtime(valor).strftime("%d/%m/%Y %H:%M")


def _aplicar_estilo_base(ws, header_row: int, total_columns: int) -> None:
    header_fill = PatternFill("solid", fgColor="1B5E3B")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D4E4D4"))
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(header_row, total_columns).coordinate}"


def _ajustar_anchos(ws, anchos: list[int]) -> None:
    for index, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = ancho


def _agregar_encabezado(ws, titulo: str, subtitulo: str, filtros: dict[str, str] | None = None) -> int:
    ws["A1"] = "ListaEsperaCCR"
    ws["A1"].font = Font(bold=True, size=16, color="1B5E3B")
    ws["A2"] = titulo
    ws["A2"].font = Font(bold=True, size=13)
    ws["A3"] = subtitulo
    ws["A4"] = f"Fecha de generación: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
    if filtros:
        texto_filtros = ", ".join(f"{clave}: {valor}" for clave, valor in filtros.items() if valor)
        texto_filtros = sanitizar_celda_excel(texto_filtros or "Sin filtros")
        ws["A5"] = f"Filtros aplicados: {texto_filtros or 'Sin filtros'}"
        return 7
    ws["A5"] = "Filtros aplicados: Sin filtros"
    return 7


def crear_excel_pacientes(
    pacientes,
    *,
    titulo: str,
    subtitulo: str,
    filtros: dict[str, str] | None = None,
    incluir_importacion: bool = False,
    periodo: str = "",
    mensaje_vacio: str = "Sin pacientes para mostrar.",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    header_row = _agregar_encabezado(ws, titulo, subtitulo, filtros)

    columnas = [
        "ID CCR",
        "Fecha derivación",
        "Días en lista",
        "Nombre",
        "RUT",
        "Edad",
        "Desde",
        "Diagnóstico",
        "Profesional",
        "Prioridad",
        "Categoría",
        "Responsable CCR",
        "Estado",
        "Contactos",
        "Inasistencias",
        "Próxima atención",
        "Fecha ingreso",
        "Fecha egreso",
        "Observaciones operativas",
    ]
    if incluir_importacion:
        columnas.extend(["Importación origen", "Periodo", "Fecha subida importación"])

    ws.append([])
    ws.append(columnas)
    _aplicar_estilo_base(ws, header_row, len(columnas))

    pacientes = list(pacientes)
    if not pacientes:
        ws.append([mensaje_vacio])
    else:
        hoy = timezone.localdate()
        for paciente in pacientes:
            importacion = getattr(paciente, "importacion_origen", None)
            fila = [
                paciente.id_ccr,
                _fecha(paciente.fecha_derivacion),
                _dias_en_lista(paciente, hoy),
                paciente.nombre,
                paciente.rut,
                paciente.edad,
                paciente.percapita_desde,
                paciente.diagnostico,
                paciente.profesional,
                paciente.get_prioridad_display(),
                paciente.get_categoria_display(),
                paciente.kine_asignado.nombre if paciente.kine_asignado else "",
                paciente.get_estado_display(),
                paciente.n_intentos_contacto,
                paciente.n_inasistencias,
                _fecha_hora(paciente.proxima_atencion),
                _fecha(paciente.fecha_ingreso),
                _fecha(paciente.fecha_egreso),
                paciente.observaciones,
            ]
            if incluir_importacion:
                fila.extend(
                    [
                        importacion.archivo_nombre if importacion else "",
                        periodo,
                        _fecha_hora(importacion.fecha_subida) if importacion else "",
                    ]
                )
            ws.append([sanitizar_celda_excel(valor) for valor in fila])
            row_number = ws.max_row
            fill_color = PRIORIDAD_FILL.get(paciente.prioridad)
            if fill_color:
                ws.cell(row=row_number, column=10).fill = PatternFill("solid", fgColor=fill_color)

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _ajustar_anchos(
        ws,
        [13, 16, 13, 28, 14, 8, 16, 34, 20, 16, 18, 22, 16, 16, 14, 18, 14, 14, 36, 24, 18, 22],
    )
    return wb


def fecha_archivo_hoy() -> str:
    return timezone.localdate().strftime("%Y%m%d")
