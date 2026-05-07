from datetime import timedelta
from io import BytesIO

from openpyxl import load_workbook
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from apps.usuarios.models import Usuario

from .models import (
    InasistenciaPaciente,
    LlamadoPaciente,
    MovimientoPaciente,
    Paciente,
    RegistroAgendaPaciente,
)


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class PacienteWorkflowTests(APITestCase):
    def setUp(self):
        self.kine = Usuario.objects.create_user(
            rut="11111111-1",
            password="testpass",
            nombre="Kine Test",
            rol=Usuario.Rol.KINE,
        )
        self.admin = Usuario.objects.create_user(
            rut="22222222-2",
            password="testpass",
            nombre="Admin Test",
            rol=Usuario.Rol.ADMIN,
        )
        self.client.force_authenticate(self.admin)
        self.paciente_counter = 0

    def crear_paciente(self, **overrides):
        self.paciente_counter += 1
        data = {
            "fecha_derivacion": timezone.localdate(),
            "percapita_desde": "CESFAM",
            "nombre": "Paciente Test",
            "rut": f"900000{self.paciente_counter:02d}-{self.paciente_counter % 10}",
            "edad": 45,
            "diagnostico": "Lumbago",
            "profesional": "KINESIOLOGO",
            "prioridad": Paciente.Prioridad.MODERADA,
            "categoria": Paciente.Categoria.LUMBAGOS,
            "kine_asignado": self.kine,
        }
        data.update(overrides)
        return Paciente.objects.create(**data)

    def test_listado_incluye_aliases_de_responsable_y_campos_compatibles(self):
        paciente = self.crear_paciente(nombre="Paciente Alias")

        response = self.client.get("/api/pacientes/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data.get("results", response.data) if isinstance(response.data, dict) else response.data
        item = next(p for p in data if p["id"] == paciente.id)
        self.assertEqual(item["kine_asignado"], self.kine.id)
        self.assertEqual(item["kine_asignado_nombre"], self.kine.nombre)
        self.assertEqual(item["responsable_asignado"], self.kine.id)
        self.assertEqual(item["responsable_nombre"], self.kine.nombre)

    def test_categoria_borrador_mantiene_valor_y_muestra_no_categorizado(self):
        paciente = self.crear_paciente(categoria=Paciente.Categoria.BORRADOR)

        response = self.client.get("/api/pacientes/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.data.get("results", response.data) if isinstance(response.data, dict) else response.data
        item = next(p for p in data if p["id"] == paciente.id)
        self.assertEqual(item["categoria"], Paciente.Categoria.BORRADOR)
        self.assertEqual(item["categoria_label"], "No categorizado")

    def test_registrar_llamado_contesto_crea_historial_e_ingresa(self):
        paciente = self.crear_paciente(fecha_ingreso=None)

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-llamado/",
            {"contesto": True, "notas": "Confirma asistencia"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.INGRESADO)
        self.assertEqual(paciente.n_intentos_contacto, 0)
        self.assertIsNotNone(paciente.fecha_ingreso)
        self.assertEqual(LlamadoPaciente.objects.count(), 1)
        self.assertEqual(
            LlamadoPaciente.objects.first().resultado,
            LlamadoPaciente.Resultado.CONTESTA_CONFIRMADO,
        )

    def test_registrar_llamado_no_contesta_desde_pendiente_pasa_a_rescate(self):
        paciente = self.crear_paciente()

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-llamado/",
            {"contesto": False, "notas": ""},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.n_intentos_contacto, 1)
        self.assertEqual(paciente.estado, Paciente.Estado.RESCATE)
        self.assertNotEqual(paciente.estado, Paciente.Estado.ABANDONO)
        self.assertEqual(LlamadoPaciente.objects.filter(paciente=paciente).count(), 1)
        self.assertEqual(MovimientoPaciente.objects.filter(paciente=paciente).count(), 1)
        self.assertEqual(
            MovimientoPaciente.objects.get(paciente=paciente).notas,
            "Primer contacto sin respuesta. Pasa a RESCATE.",
        )

    def test_registrar_llamado_rescate_sin_observacion_no_egresa(self):
        paciente = self.crear_paciente(
            estado=Paciente.Estado.RESCATE,
            n_intentos_contacto=1,
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-llamado/",
            {"contesto": False, "notas": ""},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.RESCATE)
        self.assertEqual(paciente.n_intentos_contacto, 1)
        self.assertEqual(LlamadoPaciente.objects.filter(paciente=paciente).count(), 0)

    def test_registrar_llamado_rescate_con_observacion_egresa_administrativamente(self):
        paciente = self.crear_paciente(
            estado=Paciente.Estado.RESCATE,
            n_intentos_contacto=1,
            fecha_egreso=None,
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-llamado/",
            {"contesto": False, "notas": "Segundo contacto sin respuesta"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.n_intentos_contacto, 2)
        self.assertEqual(paciente.estado, Paciente.Estado.EGRESO_ADMINISTRATIVO)
        self.assertIsNotNone(paciente.fecha_egreso)
        self.assertNotEqual(paciente.estado, Paciente.Estado.ABANDONO)
        self.assertEqual(LlamadoPaciente.objects.filter(paciente=paciente).count(), 1)
        self.assertEqual(MovimientoPaciente.objects.filter(paciente=paciente).count(), 1)

    def test_registrar_llamado_rechaza_paciente_ingresado(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO)

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-llamado/",
            {"contesto": False, "notas": "No corresponde"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.INGRESADO)

    def test_cambiar_estado_bloquea_abandono_antes_de_ingreso(self):
        for estado_actual in [Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE]:
            paciente = self.crear_paciente(estado=estado_actual, rut=f"1234567{estado_actual[0]}-5")
            response = self.client.post(
                f"/api/pacientes/{paciente.id}/cambiar-estado/",
                {"estado": Paciente.Estado.ABANDONO, "notas": "Cierre"},
                format="json",
            )
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            paciente.refresh_from_db()
            self.assertEqual(paciente.estado, estado_actual)

    def test_cambiar_estado_finales_requieren_nota_y_setean_fecha_egreso(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO, fecha_egreso=None)

        sin_nota = self.client.post(
            f"/api/pacientes/{paciente.id}/cambiar-estado/",
            {"estado": Paciente.Estado.ABANDONO, "notas": ""},
            format="json",
        )
        self.assertEqual(sin_nota.status_code, status.HTTP_400_BAD_REQUEST)

        abandono = self.client.post(
            f"/api/pacientes/{paciente.id}/cambiar-estado/",
            {"estado": Paciente.Estado.ABANDONO, "notas": "Dos inasistencias evaluadas"},
            format="json",
        )
        self.assertEqual(abandono.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.ABANDONO)
        self.assertIsNotNone(paciente.fecha_egreso)

    def test_cambiar_estado_derivado_permitido_desde_ingresado_con_nota(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO)

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/cambiar-estado/",
            {"estado": Paciente.Estado.DERIVADO, "notas": "Derivado a otro dispositivo"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.DERIVADO)
        self.assertIsNotNone(paciente.fecha_egreso)

    def test_registrar_inasistencia_reglas_y_alerta(self):
        pendiente = self.crear_paciente()
        rechazado = self.client.post(
            f"/api/pacientes/{pendiente.id}/registrar-inasistencia/",
            {
                "fecha": str(timezone.localdate()),
                "justificada": False,
                "motivo": "No asiste",
            },
            format="json",
        )
        self.assertEqual(rechazado.status_code, status.HTTP_400_BAD_REQUEST)

        ingresado = self.crear_paciente(estado=Paciente.Estado.INGRESADO, rut="87654321-9")
        justificada = self.client.post(
            f"/api/pacientes/{ingresado.id}/registrar-inasistencia/",
            {
                "fecha": str(timezone.localdate()),
                "justificada": True,
                "motivo": "Aviso previo",
            },
            format="json",
        )
        self.assertEqual(justificada.status_code, status.HTTP_201_CREATED)
        ingresado.refresh_from_db()
        self.assertEqual(ingresado.n_inasistencias, 0)

        for index in range(2):
            response = self.client.post(
                f"/api/pacientes/{ingresado.id}/registrar-inasistencia/",
                {
                    "fecha": str(timezone.localdate()),
                    "justificada": False,
                    "motivo": f"No asiste {index + 1}",
                },
                format="json",
            )
            self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        ingresado.refresh_from_db()
        self.assertEqual(ingresado.n_inasistencias, 2)
        self.assertTrue(response.data["alerta_abandono"])
        self.assertEqual(ingresado.estado, Paciente.Estado.INGRESADO)
        self.assertEqual(InasistenciaPaciente.objects.filter(paciente=ingresado).count(), 3)

    def test_registrar_asistencia_ingresa_pendiente_y_limpia_agenda(self):
        fecha = timezone.now() + timedelta(days=1)
        paciente = self.crear_paciente(
            estado=Paciente.Estado.PENDIENTE,
            fecha_ingreso=None,
            proxima_atencion=fecha,
            fecha_siguiente_cita=timezone.localdate(fecha),
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-asistencia/",
            {"fecha_programada": fecha.isoformat(), "observacion": "Asiste"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.INGRESADO)
        self.assertIsNotNone(paciente.fecha_ingreso)
        self.assertIsNone(paciente.proxima_atencion)
        self.assertIsNone(paciente.fecha_siguiente_cita)
        self.assertEqual(
            RegistroAgendaPaciente.objects.get(paciente=paciente).resultado,
            RegistroAgendaPaciente.Resultado.ASISTIO,
        )
        self.assertTrue(
            MovimientoPaciente.objects.filter(
                paciente=paciente,
                notas="Paciente asistió a atención programada.",
            ).exists()
        )

    def test_registrar_asistencia_ingresado_mantiene_estado_y_limpia_agenda(self):
        fecha = timezone.now() + timedelta(days=1)
        paciente = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            proxima_atencion=fecha,
            fecha_siguiente_cita=timezone.localdate(fecha),
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-asistencia/",
            {"fecha_programada": fecha.isoformat()},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.INGRESADO)
        self.assertIsNone(paciente.proxima_atencion)
        self.assertEqual(
            RegistroAgendaPaciente.objects.filter(
                paciente=paciente,
                resultado=RegistroAgendaPaciente.Resultado.ASISTIO,
            ).count(),
            1,
        )

    def test_registrar_inasistencia_agenda_incrementa_y_alerta_sin_abandono(self):
        fecha = timezone.now() + timedelta(days=1)
        paciente = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            n_inasistencias=1,
            proxima_atencion=fecha,
            fecha_siguiente_cita=timezone.localdate(fecha),
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/registrar-inasistencia-agenda/",
            {
                "fecha_programada": fecha.isoformat(),
                "motivo": "No asiste a atención programada.",
                "justificada": False,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        paciente.refresh_from_db()
        self.assertEqual(paciente.n_inasistencias, 2)
        self.assertEqual(paciente.estado, Paciente.Estado.INGRESADO)
        self.assertIsNone(paciente.proxima_atencion)
        self.assertTrue(response.data["alerta_abandono"])
        self.assertEqual(InasistenciaPaciente.objects.filter(paciente=paciente).count(), 1)
        self.assertEqual(
            RegistroAgendaPaciente.objects.get(paciente=paciente).resultado,
            RegistroAgendaPaciente.Resultado.NO_ASISTIO,
        )
        self.assertTrue(
            MovimientoPaciente.objects.filter(
                paciente=paciente,
                notas="No asiste a atención programada.",
            ).exists()
        )

    def test_reagendar_atencion_actualiza_proxima_y_no_cambia_estado(self):
        fecha = timezone.now() + timedelta(days=1)
        nueva_fecha = fecha + timedelta(days=3)
        paciente = self.crear_paciente(
            estado=Paciente.Estado.RESCATE,
            proxima_atencion=fecha,
            fecha_siguiente_cita=timezone.localdate(fecha),
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/reagendar-atencion/",
            {
                "fecha_programada": fecha.isoformat(),
                "nueva_fecha": nueva_fecha.isoformat(),
                "observacion": "Solicitud del paciente.",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertEqual(paciente.estado, Paciente.Estado.RESCATE)
        self.assertEqual(paciente.proxima_atencion, nueva_fecha)
        self.assertEqual(paciente.fecha_siguiente_cita, timezone.localdate(nueva_fecha))
        registro = RegistroAgendaPaciente.objects.get(paciente=paciente)
        self.assertEqual(registro.resultado, RegistroAgendaPaciente.Resultado.REAGENDADO)
        self.assertEqual(registro.nueva_fecha, nueva_fecha)

    def test_eliminar_cita_limpia_proxima_y_no_borra_paciente(self):
        fecha = timezone.now() + timedelta(days=1)
        paciente = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            proxima_atencion=fecha,
            fecha_siguiente_cita=timezone.localdate(fecha),
        )

        response = self.client.post(
            f"/api/pacientes/{paciente.id}/eliminar-cita/",
            {"fecha_programada": fecha.isoformat(), "observacion": "Cita eliminada."},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        paciente.refresh_from_db()
        self.assertIsNone(paciente.proxima_atencion)
        self.assertTrue(Paciente.objects.filter(pk=paciente.pk).exists())
        self.assertEqual(
            RegistroAgendaPaciente.objects.get(paciente=paciente).resultado,
            RegistroAgendaPaciente.Resultado.CANCELADO,
        )

    def test_acciones_de_agenda_requieren_proxima_atencion(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO, proxima_atencion=None)

        endpoints = [
            ("registrar-asistencia", {}),
            ("registrar-inasistencia-agenda", {"motivo": "No asiste"}),
            ("eliminar-cita", {}),
        ]
        for endpoint, payload in endpoints:
            response = self.client.post(
                f"/api/pacientes/{paciente.id}/{endpoint}/",
                payload,
                format="json",
            )
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_historial_completo_retorna_componentes_operativos(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO)
        MovimientoPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            estado_anterior=Paciente.Estado.PENDIENTE,
            estado_nuevo=Paciente.Estado.INGRESADO,
            notas="Ingreso",
        )
        LlamadoPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            resultado=LlamadoPaciente.Resultado.CONTESTA_CONFIRMADO,
            notas="Confirma",
        )
        InasistenciaPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            fecha=timezone.localdate(),
            motivo="No asiste",
        )
        RegistroAgendaPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            fecha_programada=timezone.now(),
            resultado=RegistroAgendaPaciente.Resultado.ASISTIO,
        )

        response = self.client.get(f"/api/pacientes/{paciente.id}/historial-completo/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("paciente", response.data)
        self.assertEqual(len(response.data["movimientos"]), 1)
        self.assertEqual(len(response.data["llamados"]), 1)
        self.assertEqual(len(response.data["inasistencias"]), 1)
        self.assertEqual(len(response.data["registros_agenda"]), 1)

    def test_historial_acciones_retorna_eventos_combinados(self):
        paciente = self.crear_paciente(estado=Paciente.Estado.INGRESADO)
        MovimientoPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            estado_anterior=Paciente.Estado.PENDIENTE,
            estado_nuevo=Paciente.Estado.INGRESADO,
            notas="Ingreso",
        )
        LlamadoPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            resultado=LlamadoPaciente.Resultado.CONTESTA_CONFIRMADO,
            notas="Confirma",
        )
        InasistenciaPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            fecha=timezone.localdate(),
            motivo="No asiste",
        )
        RegistroAgendaPaciente.objects.create(
            paciente=paciente,
            usuario=self.admin,
            fecha_programada=timezone.now(),
            resultado=RegistroAgendaPaciente.Resultado.ASISTIO,
            observacion="Asiste",
        )

        response = self.client.get(f"/api/pacientes/{paciente.id}/historial-acciones/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("paciente", response.data)
        tipos = {accion["tipo"] for accion in response.data["acciones"]}
        self.assertEqual(tipos, {"CAMBIO_ESTADO", "CONTACTO", "INASISTENCIA", "AGENDA_ASISTIO"})

    def test_alertas_operativas_retorna_grupos_esperados(self):
        alta = self.crear_paciente(
            prioridad=Paciente.Prioridad.ALTA,
            kine_asignado=None,
            estado=Paciente.Estado.PENDIENTE,
        )
        antiguo = self.crear_paciente(
            fecha_derivacion=timezone.localdate() - timedelta(days=91),
            estado=Paciente.Estado.PENDIENTE,
        )
        intento = self.crear_paciente(
            estado=Paciente.Estado.PENDIENTE,
            n_intentos_contacto=1,
        )
        rescate = self.crear_paciente(estado=Paciente.Estado.RESCATE)
        sin_agenda = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            proxima_atencion=None,
        )
        abandono = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            n_inasistencias=2,
        )
        sin_telefonos = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            telefono="",
            telefono_recados="",
        )

        response = self.client.get("/api/pacientes/alertas-operativas/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertInPaciente(response.data["alta_sin_responsable"], alta)
        self.assertInPaciente(response.data["sobre_90_dias"], antiguo)
        self.assertInPaciente(response.data["pendientes_con_1_intento"], intento)
        self.assertInPaciente(response.data["rescates_activos"], rescate)
        self.assertInPaciente(
            response.data["ingresados_sin_proxima_atencion"], sin_agenda
        )
        self.assertInPaciente(response.data["posible_abandono"], abandono)
        self.assertInPaciente(response.data["telefonos_incompletos"], sin_telefonos)

    def test_alertas_operativas_limita_cada_grupo_a_8_pacientes(self):
        for _ in range(10):
            self.crear_paciente(estado=Paciente.Estado.RESCATE)

        response = self.client.get("/api/pacientes/alertas-operativas/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["rescates_activos"]["total"], 10)
        self.assertLessEqual(len(response.data["rescates_activos"]["pacientes"]), 8)

    def test_dashboard_resumen_admin_devuelve_contadores_agregados(self):
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE)
        self.crear_paciente(estado=Paciente.Estado.RESCATE)
        self.crear_paciente(estado=Paciente.Estado.INGRESADO)
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE, kine_asignado=None)
        self.crear_paciente(estado=Paciente.Estado.ALTA_MEDICA)
        self.crear_paciente(estado=Paciente.Estado.ALTA_MEDICA, kine_asignado=None)

        response = self.client.get("/api/pacientes/dashboard-resumen/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["lista_activa"], 4)
        self.assertEqual(response.data["pendientes"], 2)
        self.assertEqual(response.data["rescate"], 1)
        self.assertEqual(response.data["ingresados"], 1)
        self.assertEqual(response.data["sin_asignar"], 2)
        self.assertEqual(response.data["asignados_activos"], 3)
        self.assertEqual(response.data["mios_activos"], 0)
        self.assertEqual(response.data["rescates_globales"], 1)
        self.assertEqual(response.data["cola_llamados"], 2)

    def test_dashboard_resumen_kine_acota_resumen_y_cola_a_su_cartera(self):
        otro_kine = Usuario.objects.create_user(
            rut="33333333-3",
            password="testpass",
            nombre="Otro Kine",
            rol=Usuario.Rol.KINE,
        )
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE)
        self.crear_paciente(estado=Paciente.Estado.RESCATE)
        self.crear_paciente(estado=Paciente.Estado.INGRESADO)
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE, kine_asignado=otro_kine)
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE, kine_asignado=None)
        self.crear_paciente(estado=Paciente.Estado.DERIVADO)
        self.client.force_authenticate(self.kine)

        response = self.client.get("/api/pacientes/dashboard-resumen/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["lista_activa"], 3)
        self.assertEqual(response.data["pendientes"], 1)
        self.assertEqual(response.data["rescate"], 1)
        self.assertEqual(response.data["ingresados"], 1)
        self.assertEqual(response.data["sin_asignar"], 1)
        self.assertEqual(response.data["asignados_activos"], 4)
        self.assertEqual(response.data["mios_activos"], 3)
        self.assertEqual(response.data["rescates_globales"], 1)
        self.assertEqual(response.data["cola_llamados"], 2)

    def assertInPaciente(self, grupo, paciente):
        ids = {item["id"] for item in grupo["pacientes"]}
        self.assertIn(paciente.id, ids)

    def test_filtro_alerta_alta_sin_responsable(self):
        esperado = self.crear_paciente(
            prioridad=Paciente.Prioridad.ALTA,
            kine_asignado=None,
            estado=Paciente.Estado.PENDIENTE,
        )
        self.crear_paciente(
            prioridad=Paciente.Prioridad.ALTA,
            estado=Paciente.Estado.PENDIENTE,
        )

        response = self.client.get("/api/pacientes/?alerta=alta_sin_responsable")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_sobre_90_dias(self):
        esperado = self.crear_paciente(
            fecha_derivacion=timezone.localdate() - timedelta(days=91),
            estado=Paciente.Estado.RESCATE,
        )
        self.crear_paciente(
            fecha_derivacion=timezone.localdate() - timedelta(days=91),
            estado=Paciente.Estado.INGRESADO,
        )
        self.crear_paciente(
            fecha_derivacion=timezone.localdate() - timedelta(days=20),
            estado=Paciente.Estado.PENDIENTE,
        )

        response = self.client.get("/api/pacientes/?alerta=sobre_90_dias")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_pendientes_con_1_intento(self):
        esperado = self.crear_paciente(
            estado=Paciente.Estado.PENDIENTE,
            n_intentos_contacto=1,
        )
        self.crear_paciente(
            estado=Paciente.Estado.PENDIENTE,
            n_intentos_contacto=2,
        )

        response = self.client.get("/api/pacientes/?alerta=pendientes_con_1_intento")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_rescates_activos(self):
        esperado = self.crear_paciente(estado=Paciente.Estado.RESCATE)
        self.crear_paciente(estado=Paciente.Estado.PENDIENTE)

        response = self.client.get("/api/pacientes/?alerta=rescates_activos")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_ingresados_sin_proxima_atencion(self):
        esperado = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            proxima_atencion=None,
        )
        self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            proxima_atencion=timezone.now(),
        )

        response = self.client.get(
            "/api/pacientes/?alerta=ingresados_sin_proxima_atencion"
        )

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_posible_abandono(self):
        esperado = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            n_inasistencias=2,
        )
        self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            n_inasistencias=1,
        )

        response = self.client.get("/api/pacientes/?alerta=posible_abandono")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_telefonos_incompletos(self):
        esperado = self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            telefono="",
            telefono_recados="",
        )
        self.crear_paciente(
            estado=Paciente.Estado.INGRESADO,
            telefono="123456789",
            telefono_recados="",
        )
        self.crear_paciente(
            estado=Paciente.Estado.ALTA_MEDICA,
            telefono="",
            telefono_recados="",
        )

        response = self.client.get("/api/pacientes/?alerta=telefonos_incompletos")

        self.assertResponseSoloPacientes(response, [esperado])

    def test_filtro_alerta_desconocida_no_rompe(self):
        paciente = self.crear_paciente()

        response = self.client.get("/api/pacientes/?alerta=no_existe")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertResponseSoloPacientes(response, [paciente])

    def assertResponseSoloPacientes(self, response, pacientes):
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        expected_ids = {paciente.id for paciente in pacientes}
        response_ids = {item["id"] for item in response.data}
        self.assertEqual(response_ids, expected_ids)

    def test_exportar_lista_espera_devuelve_excel(self):
        self.crear_paciente(nombre="Paciente Exportado")

        response = self.client.get("/api/pacientes/exportar/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], EXCEL_CONTENT_TYPE)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        self.assertEqual(ws["A1"].value, "ListaEsperaCCR")
        self.assertIn("ID CCR", [cell.value for cell in ws[7]])

    def test_exportar_lista_espera_respeta_alerta_rescates(self):
        self.crear_paciente(nombre="Paciente Pendiente", estado=Paciente.Estado.PENDIENTE)
        rescate = self.crear_paciente(nombre="Paciente Rescate", estado=Paciente.Estado.RESCATE)

        response = self.client.get("/api/pacientes/exportar/?alerta=rescates_activos")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        nombres = [row[3] for row in ws.iter_rows(min_row=8, values_only=True) if row[3]]
        self.assertEqual(nombres, [rescate.nombre])

    def test_exportar_lista_espera_sanitiza_formula_excel(self):
        self.crear_paciente(nombre="=cmd", diagnostico="+diagnostico", observaciones="@observacion")

        response = self.client.get("/api/pacientes/exportar/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        fila = next(row for row in ws.iter_rows(min_row=8, values_only=True) if row[3])
        self.assertEqual(fila[3], "'=cmd")
        self.assertEqual(fila[7], "'+diagnostico")
        self.assertEqual(fila[18], "'@observacion")

    def test_exportar_lista_espera_muestra_no_categorizado(self):
        self.crear_paciente(categoria=Paciente.Categoria.BORRADOR)

        response = self.client.get("/api/pacientes/exportar/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook.active
        fila = next(row for row in ws.iter_rows(min_row=8, values_only=True) if row[3])
        self.assertEqual(fila[10], "No categorizado")

    def test_usuario_no_autenticado_no_puede_exportar(self):
        self.client.force_authenticate(user=None)

        response = self.client.get("/api/pacientes/exportar/")

        self.assertIn(
            response.status_code,
            {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN},
        )
